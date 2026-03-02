#!/usr/bin/env python3
"""
Phase 6: Platform/Module Deep Import
Parses P-PlatformData sheet from Kamstrup Excel and creates ModuleFactSheet JSON files.

Each row = a module within a platform, mapped to a business capability.
Links modules to parent Platform (ITComponentFactSheet) and BusinessCapability.
"""

import json
import uuid
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────
EXCEL_PATH = r"F:\RootContext\Kamstrup\Kamstrup_Business-Capability-Map.xlsx"
LOOKUPS_PATH = r"F:\RootContext\Kamstrup\import-scripts\output\lookups-phase4.json"
OUTPUT_DIR = r"F:\RootContext\Kamstrup\import-scripts\output\modules"
SHEET_NAME = "P-PlatformData"
HEADER_ROW = 20
DATA_START_ROW = 21

# ── Deterministic GUID ────────────────────────────────────────────────────
NAMESPACE_MODULES = uuid.UUID("b7e3c9a1-4f2d-4e8b-9c6a-1d3f5e7a9b0c")

def deterministic_guid(platform: str, capability_id: str, module_name: str, submodule: str) -> str:
    """Generate a deterministic UUID from composite key."""
    key = f"module:{platform}:{capability_id}:{module_name or ''}:{submodule or ''}"
    return str(uuid.uuid5(NAMESPACE_MODULES, key))


# ── Enum Mappings ──────────────────────────────────────────────────────────
UTILIZATION_MAP = {
    "No": "None",
    "Low": "Low",
    "Medium": "Medium",
    "High": "High",
    "?": "Unknown",
    None: "Unknown",
    "": "Unknown",
}

TECHNICAL_FIT_MAP = {
    "Fully Approoriate": "Appropriate",   # ⚠️ Source data typo — exact match
    "Fully Appropriate": "Appropriate",    # Just in case they fix the typo
    "Adequate": "Adequate",
    "Unreasonable": "Unreasonable",
    "Inappropriate": "Inappropriate",
    None: "Unmapped",
    "": "Unmapped",
}

FUNCTIONAL_FIT_MAP = {
    "Perfect": "Perfect",
    "Appropriate": "Appropriate",
    "Insufficient": "Insufficient",
    "Unreasonable": "Unreasonable",
    None: "Unmapped",
    "": "Unmapped",
}

# Platform name normalization (sheet -> lookup key)
PLATFORM_NAME_FIXES = {
    "JetBrain": "JetBrains",
}


def parse_bool(val) -> bool:
    """Parse Yes/No to bool."""
    if val is None:
        return False
    return str(val).strip().lower() == "yes"


def parse_date(val):
    """Parse a datetime or date value to ISO string or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%dT00:00:00")
    if isinstance(val, str):
        val = val.strip()
        if val in ("", "#N/A"):
            return None
        try:
            return datetime.strptime(val, "%Y-%m-%d").strftime("%Y-%m-%dT00:00:00")
        except ValueError:
            return None
    return None


def parse_eol(val):
    """Parse Expected End-of-Life (could be year integer, date, or #N/A)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%dT00:00:00")
    if isinstance(val, (int, float)):
        year = int(val)
        if 2000 <= year <= 2100:
            return f"{year}-12-31T00:00:00"
        return None
    if isinstance(val, str):
        val = val.strip()
        if val in ("", "#N/A"):
            return None
        try:
            year = int(val)
            if 2000 <= year <= 2100:
                return f"{year}-12-31T00:00:00"
        except ValueError:
            pass
    return None


def build_module_factsheet(row_data, row_num, platform_ids, bc_ids):
    """Build a ModuleFactSheet dict from a row."""
    platform_name_raw = str(row_data[1]).strip()  # Col B
    capability_key = str(row_data[2]).strip() if row_data[2] else None  # Col C
    module_name = str(row_data[3]).strip() if row_data[3] else None  # Col D
    submodule = str(row_data[4]).strip() if row_data[4] else None  # Col E
    utilization = row_data[5]  # Col F
    tech_fit = row_data[6]  # Col G
    func_fit = row_data[7]  # Col H
    ai_val = row_data[8]  # Col I
    description = str(row_data[9]).strip() if row_data[9] else None  # Col J
    long_desc = str(row_data[10]).strip() if row_data[10] else None  # Col K
    purchased = row_data[11]  # Col L
    from_date = row_data[12]  # Col M
    to_date = row_data[13]  # Col N
    app_ref = str(row_data[14]).strip() if row_data[14] else None  # Col O
    capability_check = str(row_data[15]).strip() if row_data[15] else None  # Col P
    eol = row_data[16]  # Col Q

    # Fix platform name
    platform_name = PLATFORM_NAME_FIXES.get(platform_name_raw, platform_name_raw)

    # Resolve platform ID
    platform_id = platform_ids.get(platform_name)
    if not platform_id:
        print(f"  WARNING row {row_num}: Platform '{platform_name_raw}' not found in lookups — skipping")
        return None

    # Resolve capability ID
    capability_id = None
    if capability_key:
        capability_id = bc_ids.get(capability_key)
        if not capability_id:
            # Try partial match on the numeric prefix
            cap_num = capability_key.split("-")[0] if "-" in capability_key else capability_key
            for k, v in bc_ids.items():
                if k.startswith(cap_num + "-"):
                    capability_id = v
                    break

    # Build display name
    if submodule and submodule not in ("#N/A", "None"):
        display_name = submodule
    elif module_name and module_name not in ("#N/A", "None", "No module"):
        display_name = module_name
    else:
        # Fallback: Platform — Capability
        cap_parts = capability_key.split("-") if capability_key else ["?"]
        cap_name = cap_parts[1] if len(cap_parts) > 1 else cap_parts[0]
        display_name = f"{platform_name} — {cap_name}"

    # Clean module_name for #N/A
    if module_name in ("#N/A", "None"):
        module_name = None

    # Generate deterministic ID
    module_id = deterministic_guid(platform_name, capability_key or "", module_name or "", submodule or "")

    # Map enums
    util_str = str(utilization).strip() if utilization else None
    util_mapped = UTILIZATION_MAP.get(util_str, "Unknown")

    tech_str = str(tech_fit).strip() if tech_fit else None
    tech_mapped = TECHNICAL_FIT_MAP.get(tech_str, "Unmapped")

    func_str = str(func_fit).strip() if func_fit else None
    func_mapped = FUNCTIONAL_FIT_MAP.get(func_str, "Unmapped")

    ai_enabled = parse_bool(ai_val)
    is_purchased = parse_bool(purchased)

    # Build lifecycle
    lifecycle = {}
    active_date = parse_date(from_date)
    phase_out = parse_date(to_date)
    end_of_life = parse_eol(eol)
    if active_date:
        lifecycle["Active"] = active_date
    if phase_out:
        lifecycle["PhaseOut"] = phase_out
    if end_of_life:
        lifecycle["EndOfLife"] = end_of_life

    # Build the factsheet
    fs = {
        "Id": module_id,
        "DisplayName": display_name,
        "ShortDescription": description,
        "RichDescription": long_desc,
        "ModuleType": "Native" if not is_purchased else "3rd Party",
        "Utilization": util_mapped,
        "ModuleTechnicalFit": tech_mapped,
        "ModuleFunctionalFit": func_mapped,
        "IsPurchased": is_purchased,
        "AIEnabled": ai_enabled,
        "FactSheetType": "ModuleFactSheet",
        # Parent platform
        "ParentPlatformId": platform_id,
        "ParentPlatformName": platform_name,
        # Capability link
        "BusinessCapabilityId": capability_id,
        "BusinessCapabilityKey": capability_key,
        # Application reference (raw)
        "ApplicationReference": app_ref if app_ref and app_ref != "None" else None,
        # Capability sub-area tag
        "CapabilitySubArea": capability_check if capability_check and capability_check not in ("0", "None", "#N/A") else None,
        # Module name from col D (for context)
        "ModuleOrApplicationName": module_name,
    }

    if lifecycle:
        fs["LifeCycle"] = lifecycle

    # Source metadata
    fs["_sourceRow"] = row_num
    fs["_sourcePlatform"] = platform_name_raw

    return fs


def main():
    print("=" * 70)
    print("Phase 6: Platform/Module Deep Import")
    print("=" * 70)

    # Load lookups
    print(f"\nLoading lookups from {LOOKUPS_PATH}...")
    with open(LOOKUPS_PATH) as f:
        lookups = json.load(f)

    platform_ids = lookups["platform_ids"]
    bc_ids = lookups["business_capability_ids"]
    print(f"  Platforms: {len(platform_ids)}, Business Capabilities: {len(bc_ids)}")

    # Load Excel
    print(f"\nLoading Excel from {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    print(f"  Sheet '{SHEET_NAME}': {ws.max_row} rows × {ws.max_column} cols")

    # Parse data rows
    modules = []
    skipped = 0
    seen_keys = {}  # For duplicate detection

    for row_num in range(DATA_START_ROW, ws.max_row + 1):
        platform = ws.cell(row=row_num, column=2).value
        if not platform:
            continue

        row_data = [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]
        fs = build_module_factsheet(row_data, row_num, platform_ids, bc_ids)

        if fs is None:
            skipped += 1
            continue

        # Check for duplicates (same ID = same composite key)
        if fs["Id"] in seen_keys:
            prev_row = seen_keys[fs["Id"]]
            print(f"  WARNING: Duplicate key at row {row_num} (same as row {prev_row}): {fs['DisplayName']}")
            # Append row number to make unique
            fs["Id"] = deterministic_guid(
                fs["ParentPlatformName"],
                fs.get("BusinessCapabilityKey", ""),
                fs.get("ModuleOrApplicationName", ""),
                f"{fs['DisplayName']}_{row_num}"
            )

        seen_keys[fs["Id"]] = row_num
        modules.append(fs)

    print(f"\n  Parsed: {len(modules)} modules, Skipped: {skipped}")

    # Statistics
    platforms_used = set(m["ParentPlatformName"] for m in modules)
    util_dist = {}
    tech_dist = {}
    func_dist = {}
    linked_caps = sum(1 for m in modules if m["BusinessCapabilityId"])
    linked_apps = sum(1 for m in modules if m["ApplicationReference"])
    purchased_count = sum(1 for m in modules if m["IsPurchased"])
    ai_count = sum(1 for m in modules if m["AIEnabled"])

    for m in modules:
        util_dist[m["Utilization"]] = util_dist.get(m["Utilization"], 0) + 1
        tech_dist[m["ModuleTechnicalFit"]] = tech_dist.get(m["ModuleTechnicalFit"], 0) + 1
        func_dist[m["ModuleFunctionalFit"]] = func_dist.get(m["ModuleFunctionalFit"], 0) + 1

    print(f"\n  Platforms: {sorted(platforms_used)}")
    print(f"  Utilization: {dict(sorted(util_dist.items()))}")
    print(f"  Technical Fit: {dict(sorted(tech_dist.items()))}")
    print(f"  Functional Fit: {dict(sorted(func_dist.items()))}")
    print(f"  Linked to capabilities: {linked_caps}/{len(modules)}")
    print(f"  Linked to applications: {linked_apps}/{len(modules)}")
    print(f"  Purchased: {purchased_count}, AI-enabled: {ai_count}")

    # Write output
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Individual module files
    for m in modules:
        safe_name = m["DisplayName"].replace("/", "_").replace("\\", "_").replace(":", "_").replace("*", "_").replace("?", "_").replace('"', "_").replace("<", "_").replace(">", "_").replace("|", "_")
        filename = f"{m['ParentPlatformName']}_{safe_name}_{m['Id'][:8]}.json"
        filepath = os.path.join(OUTPUT_DIR, filename)
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(m, f, indent=2, ensure_ascii=False)

    # Summary file with all modules
    summary_path = os.path.join(OUTPUT_DIR, "_all_modules.json")
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump({
            "count": len(modules),
            "generated": datetime.now().isoformat(),
            "platforms": sorted(platforms_used),
            "utilization_distribution": util_dist,
            "technical_fit_distribution": tech_dist,
            "functional_fit_distribution": func_dist,
            "modules": modules,
        }, f, indent=2, ensure_ascii=False)

    # Relationships file — for import into OmniGaze
    relationships = []
    for m in modules:
        # Module -> Platform (hierarchy parent)
        relationships.append({
            "type": "hierarchy",
            "parentId": m["ParentPlatformId"],
            "parentType": "ITComponentFactSheet",
            "childId": m["Id"],
            "childType": "ModuleFactSheet",
        })
        # Module -> BusinessCapability (relationship)
        if m["BusinessCapabilityId"]:
            relationships.append({
                "type": "relationship",
                "parentId": m["BusinessCapabilityId"],
                "parentType": "BusinessCapabilityFactSheet",
                "childId": m["Id"],
                "childType": "ModuleFactSheet",
            })

    rel_path = os.path.join(OUTPUT_DIR, "_relationships.json")
    with open(rel_path, "w", encoding="utf-8") as f:
        json.dump(relationships, f, indent=2, ensure_ascii=False)

    # Update lookups with module IDs
    module_ids = {}
    for m in modules:
        key = f"{m['ParentPlatformName']}|{m['DisplayName']}"
        module_ids[key] = m["Id"]

    lookups_output_path = os.path.join(os.path.dirname(OUTPUT_DIR), "lookups-phase6.json")
    with open(lookups_output_path, "w", encoding="utf-8") as f:
        json.dump({
            "module_ids": module_ids,
            "module_count": len(modules),
            "generated": datetime.now().isoformat(),
        }, f, indent=2, ensure_ascii=False)

    print(f"\n  Output written to: {OUTPUT_DIR}")
    print(f"  Individual files: {len(modules)}")
    print(f"  Summary: {summary_path}")
    print(f"  Relationships: {rel_path} ({len(relationships)} entries)")
    print(f"  Lookups: {lookups_output_path}")
    print(f"\n{'=' * 70}")
    print("Phase 6 COMPLETE")
    print(f"{'=' * 70}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
