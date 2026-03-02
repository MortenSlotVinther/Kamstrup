#!/usr/bin/env python3
"""
Phase 3: Kamstrup Application Import Script
============================================
Parses A-Applications (992 active) and A-Applications_Removed (~88 retired)
from Kamstrup_Business-Capability-Map.xlsx and produces JSON files ready
for import into OmniGaze via the MCP API or direct FactSheetContainer merge.

Output files:
  - output/applications_active.json      (992 ApplicationFactSheets)
  - output/applications_removed.json     (~88 retired ApplicationFactSheets)
  - output/providers_extracted.json       (unique providers from vendor/consultancy columns)
  - output/app_number_to_guid.json        (NoApplication → Guid mapping for Phase 4)
  - output/import_report.txt              (execution log)
  - output/custom_column_definitions.json (custom column defs to create in CustomerSettings)

Usage:
  python import_applications.py [--excel PATH] [--output-dir DIR]

Author: Ole (Subagent)
Date: 2026-02-07
"""

import json
import uuid
import hashlib
import os
import sys
import argparse
from datetime import datetime, date
from collections import Counter, OrderedDict
from typing import Optional, Any

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ─── Constants ─────────────────────────────────────────────────────────────

EXCEL_PATH = r"F:\RootContext\Kamstrup\Kamstrup_Business-Capability-Map.xlsx"
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")

# Deterministic GUID namespace for Kamstrup applications
KAMSTRUP_APP_NAMESPACE = uuid.UUID("a1b2c3d4-e5f6-7890-abcd-ef0123456789")

# ─── Mapping Tables ───────────────────────────────────────────────────────

LIFECYCLE_MAPPING = {
    "Under Evaluation":                 ("Plan",       "Under Evaluation"),
    "1-Strategic":                      ("Active",     "1-Strategic"),
    "2-Important application":          ("Active",     "2-Important"),
    "3-Kamstrup application":           ("Active",     "3-Kamstrup app"),
    "4-Saved for now":                  ("PhaseOut",   "4-Saved for now"),
    "5-Investigate":                    ("PhaseOut",   "5-Investigate"),
    "7-Potential for phase out":        ("PhaseOut",   "7-Potential phase out"),
    "8-Phase out":                      ("PhaseOut",   "8-Phase out"),
    "8-Kamstrup application-Phase out": ("PhaseOut",   "8-Kamstrup app-Phase out"),
    "9-End of Life":                    ("EndOfLife",  "9-End of Life"),
    "Not in use":                       ("EndOfLife",  "Not in use"),
}

INSTALL_TYPE_MAPPING = {
    "On Premise":               "OnPrem",
    "Cloud":                    "SaaS",
    "Third Party Hosted":       "ThirdPartyHosted",
    "Edge Computing":           "EdgeComputing",
    "Distributed Applications": "DistributedApp",
}

TIME_MAPPING = {
    "Tolerate":  "Tolerate",
    "Invest":    "Invest",
    "Migrate":   "Migrate",
    "Eliminate":  "Eliminate",
}

USER_BASE_MAPPING = {
    "0-9":     "VerySmall",
    "10-49":   "Small",
    "50-99":   "Medium",
    "100-499": "Large",
    "500-999": "VeryLarge",
    "1000->":  "Enterprise",
}

AI_MAPPING = {
    "No":                       (False, "NotApplicable"),
    "Yes":                      (True,  "NotEvaluated"),
    "Yes - Risk not evaluated": (True,  "NotEvaluated"),
    "Yes - Minimal Risk":       (True,  "Minimal"),
    "Yes - Limited Risk":       (True,  "Limited"),
    "Yes - High Risk":          (True,  "High"),
    "Yes - Unacceptable Risk":  (True,  "Unacceptable"),
    "Unknown":                  (False, "NotApplicable"),
}

SECURITY_STATUS_MAPPING = {
    "Yes": "Approved",
    "No":  "NotAssessed",
}

SECURITY_DEBT_MAPPING = {
    "Low":         "Low",
    "Medium":      "Medium",
    "High":        "High",
    "Investigate": "Investigate",
}


# ─── Helper Functions ──────────────────────────────────────────────────────

def generate_deterministic_guid(app_number: int) -> str:
    """Generate a deterministic GUID from application number for cross-referencing."""
    return str(uuid.uuid5(KAMSTRUP_APP_NAMESPACE, str(app_number)))


def safe_str(val: Any) -> Optional[str]:
    """Convert cell value to cleaned string, or None if empty."""
    if val is None:
        return None
    s = str(val).strip()
    if s == "" or s == "#REF!" or s == "N/A":
        return None
    return s


def safe_date(val: Any) -> Optional[str]:
    """Convert cell value to ISO date string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%dT00:00:00")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%dT00:00:00")
    s = str(val).strip()
    if not s:
        return None
    # Try parsing year-only
    if s.isdigit() and len(s) == 4:
        return f"{s}-01-01T00:00:00"
    return None


def safe_int(val: Any) -> Optional[int]:
    """Convert cell value to int, or None."""
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


class ImportReport:
    """Collects import statistics and log messages."""
    
    def __init__(self):
        self.lines = []
        self.counters = Counter()
        self.warnings = []
        self.errors = []
    
    def log(self, msg: str):
        self.lines.append(msg)
    
    def warn(self, msg: str):
        self.warnings.append(msg)
        self.lines.append(f"  ⚠️ WARNING: {msg}")
    
    def error(self, msg: str):
        self.errors.append(msg)
        self.lines.append(f"  ❌ ERROR: {msg}")
    
    def count(self, key: str, n: int = 1):
        self.counters[key] += n
    
    def dump(self) -> str:
        out = []
        out.append("=" * 70)
        out.append("  Kamstrup Phase 3: Application Import Report")
        out.append(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        out.append("=" * 70)
        out.append("")
        out.append("── Counters ──")
        for k, v in sorted(self.counters.items()):
            out.append(f"  {k}: {v}")
        out.append("")
        if self.warnings:
            out.append(f"── Warnings ({len(self.warnings)}) ──")
            for w in self.warnings:
                out.append(f"  ⚠️  {w}")
            out.append("")
        if self.errors:
            out.append(f"── Errors ({len(self.errors)}) ──")
            for e in self.errors:
                out.append(f"  ❌  {e}")
            out.append("")
        out.append("── Detailed Log ──")
        out.extend(self.lines)
        return "\n".join(out)


# ─── Active Applications Parser ───────────────────────────────────────────

def parse_active_applications(ws, report: ImportReport):
    """Parse A-Applications sheet (rows 4-995 = 992 data rows)."""
    
    apps = []
    app_number_to_guid = {}
    app_name_to_guid = {}
    providers_set = set()
    sub_component_refs = []  # (child_guid, parent_noapp_or_name)
    successor_refs = []       # (app_guid, replacement_name)
    
    report.log(f"\n── Parsing A-Applications ──")
    report.log(f"  Sheet range: rows 4 to {ws.max_row}, {ws.max_column} columns")
    
    row_count = 0
    for row_idx in range(4, ws.max_row + 1):
        # Col 4 = Application name (required)
        app_name = safe_str(ws.cell(row=row_idx, column=4).value)
        if not app_name:
            report.count("active_skipped_empty")
            continue
        
        row_count += 1
        
        # Col 2 = NoApplication (e.g. "1 - IFS ERP")
        no_application = safe_str(ws.cell(row=row_idx, column=2).value)
        # Col 3 = No (numeric ID)
        app_no = safe_int(ws.cell(row=row_idx, column=3).value)
        
        # Generate deterministic GUID
        if app_no is not None:
            app_guid = generate_deterministic_guid(app_no)
            app_number_to_guid[app_no] = app_guid
            if no_application:
                app_number_to_guid[no_application] = app_guid
        else:
            app_guid = str(uuid.uuid4())
            report.warn(f"Row {row_idx}: No app number for '{app_name}', using random GUID")
        
        app_name_to_guid[app_name] = app_guid
        
        # ── Build ApplicationFactSheet ──
        app = OrderedDict()
        app["Id"] = app_guid
        app["DisplayName"] = app_name
        app["_RowIndex"] = row_idx
        app["_NoApplication"] = no_application
        app["_AppNo"] = app_no
        
        # Col 5 - Application Group → Category
        app["Category"] = safe_str(ws.cell(row=row_idx, column=5).value) or ""
        
        # Col 6 - Introduction date → LifeCycle.Active
        intro_date = safe_date(ws.cell(row=row_idx, column=6).value)
        
        # Col 7 - Install type → HostedOn
        install_type_raw = safe_str(ws.cell(row=row_idx, column=7).value)
        hosting_enum = "Unmapped"
        if install_type_raw and install_type_raw in INSTALL_TYPE_MAPPING:
            hosting_enum = INSTALL_TYPE_MAPPING[install_type_raw]
            report.count(f"install_type_{hosting_enum}")
        elif install_type_raw:
            report.warn(f"Row {row_idx}: Unknown install type '{install_type_raw}'")
        
        app["HostedOn"] = {
            "Criticality": hosting_enum,
            "HostingTypeValue": hosting_enum if hosting_enum != "Unmapped" else None,
            "HostingTypeDescription": install_type_raw,
        }
        
        # Col 8 - Gartner TIME → PortFolioStrategy.TIME
        time_raw = safe_str(ws.cell(row=row_idx, column=8).value)
        time_enum = "Unmapped"
        if time_raw and time_raw in TIME_MAPPING:
            time_enum = TIME_MAPPING[time_raw]
            report.count(f"TIME_{time_enum}")
        elif time_raw:
            report.warn(f"Row {row_idx}: Unknown TIME value '{time_raw}'")
        
        app["PortFolioStrategy"] = {
            "TIME": time_enum,
            "Description": None,
        }
        
        # Col 9 - Application Type → ApplicationTypeTags
        app_type = safe_str(ws.cell(row=row_idx, column=9).value)
        app["ApplicationTypeTags"] = [app_type] if app_type else []
        
        # Col 10 - Software Vendor → Provider (stored for Phase 4 relationship)
        vendor = safe_str(ws.cell(row=row_idx, column=10).value)
        app["_SoftwareVendor"] = vendor
        if vendor:
            providers_set.add(vendor)
        
        # Col 11 - ItemNumber
        app["_ItemNumber"] = safe_str(ws.cell(row=row_idx, column=11).value)
        
        # Col 12 - Consultancy → Provider (stored for Phase 4)
        consultancy = safe_str(ws.cell(row=row_idx, column=12).value)
        app["_Consultancy"] = consultancy
        if consultancy:
            providers_set.add(consultancy)
        
        # Col 13 - User base → UserBase.Size
        user_base_raw = safe_str(ws.cell(row=row_idx, column=13).value)
        user_base_enum = "Unknown"
        if user_base_raw and user_base_raw in USER_BASE_MAPPING:
            user_base_enum = USER_BASE_MAPPING[user_base_raw]
            report.count(f"userbase_{user_base_enum}")
        elif user_base_raw:
            report.warn(f"Row {row_idx}: Unknown user base '{user_base_raw}'")
        
        app["UserBase"] = {
            "Size": user_base_enum,
            "Count": None,
            "Description": user_base_raw,
            "UserType": "Unknown",
        }
        
        # Col 14 - Lifecycle stage → LifeCycle + LifecycleStage
        lifecycle_raw = safe_str(ws.cell(row=row_idx, column=14).value)
        lc_phase = None
        lc_label = None
        if lifecycle_raw and lifecycle_raw in LIFECYCLE_MAPPING:
            lc_phase, lc_label = LIFECYCLE_MAPPING[lifecycle_raw]
            report.count(f"lifecycle_{lc_phase}")
        elif lifecycle_raw:
            report.warn(f"Row {row_idx}: Unknown lifecycle stage '{lifecycle_raw}'")
        
        # Build LifeCycle object
        lifecycle = {
            "Plan": None,
            "PhaseIn": None,
            "Active": None,
            "PhaseOut": None,
            "EndOfLife": None,
        }
        
        # Set intro date to the appropriate phase
        if lc_phase == "Plan" and intro_date:
            lifecycle["Plan"] = intro_date
        elif lc_phase == "Active" and intro_date:
            lifecycle["Active"] = intro_date
        elif lc_phase == "PhaseOut" and intro_date:
            lifecycle["Active"] = intro_date  # Was active from this date
            lifecycle["PhaseOut"] = intro_date  # Now phasing out
        elif lc_phase == "EndOfLife" and intro_date:
            lifecycle["Active"] = intro_date
        elif intro_date:
            lifecycle["Active"] = intro_date  # Default: treat as active date
        
        # Col 15 - Expected End-of-Life Year
        eol_date = safe_date(ws.cell(row=row_idx, column=15).value)
        if eol_date:
            lifecycle["EndOfLife"] = eol_date
        
        app["LifeCycle"] = lifecycle
        app["LifecycleStage"] = lc_label
        
        # Col 16 - Replaced by → SuccessorId (resolve in second pass)
        replaced_by = safe_str(ws.cell(row=row_idx, column=16).value)
        if replaced_by:
            successor_refs.append((app_guid, replaced_by))
        app["SuccessorId"] = []
        
        # Col 17 - Details → RichDescription
        details1 = safe_str(ws.cell(row=row_idx, column=17).value)
        # Col 21 - Details2 → append to RichDescription
        details2 = safe_str(ws.cell(row=row_idx, column=21).value)
        
        description_parts = []
        if details1:
            description_parts.append(details1)
        if details2:
            description_parts.append(details2)
        app["RichDescription"] = "\n\n".join(description_parts) if description_parts else None
        
        # Col 18 - URL
        app["Url"] = safe_str(ws.cell(row=row_idx, column=18).value)
        
        # Col 19 - Business Owner → Owners (OwnerRole.BusinessOwner)
        business_owner = safe_str(ws.cell(row=row_idx, column=19).value)
        owners = []
        if business_owner:
            owners.append({
                "Role": "BusinessOwner",
                "ExternalName": business_owner,
                "AssignedDate": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            })
        
        # Col 22 - Owner → Responsible
        owner = safe_str(ws.cell(row=row_idx, column=22).value)
        app["Responsible"] = owner
        
        if owner:
            owners.append({
                "Role": "SystemOwner",
                "ExternalName": owner,
                "AssignedDate": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            })
        
        app["Owners"] = owners
        
        # Col 20 - Supported by Group IT → OrganizationTags
        org_tags = []
        supported_it = safe_str(ws.cell(row=row_idx, column=20).value)
        if supported_it and supported_it.upper() in ("X", "YES"):
            org_tags.append("Supported by Group IT")
        
        # Col 42 - Crown Jewels → OrganizationTags
        crown_jewel = safe_str(ws.cell(row=row_idx, column=42).value)
        if crown_jewel:
            org_tags.append("CrownJewel")
        
        app["OrganizationTags"] = org_tags
        
        # Col 23 - Super User
        app["_SuperUser"] = safe_str(ws.cell(row=row_idx, column=23).value)
        
        # Col 24 - Named users
        app["_NamedUsers"] = safe_str(ws.cell(row=row_idx, column=24).value)
        
        # Col 29 - Sub component to → HierarchyParentId (resolve in second pass)
        sub_component_to = safe_str(ws.cell(row=row_idx, column=29).value)
        if sub_component_to:
            sub_component_refs.append((app_guid, sub_component_to, app_name))
        
        # Col 30 - Supporting applications (stored for Phase 4)
        app["_SupportingApplications"] = safe_str(ws.cell(row=row_idx, column=30).value)
        
        # Col 31 - Org
        app["_Org"] = safe_str(ws.cell(row=row_idx, column=31).value)
        
        # Col 32 - AI → AIClassification
        ai_raw = safe_str(ws.cell(row=row_idx, column=32).value)
        uses_ai = False
        risk_level = "NotApplicable"
        if ai_raw and ai_raw in AI_MAPPING:
            uses_ai, risk_level = AI_MAPPING[ai_raw]
            report.count(f"ai_{risk_level}")
        elif ai_raw:
            report.warn(f"Row {row_idx}: Unknown AI value '{ai_raw}'")
        
        app["AIClassification"] = {
            "UsesAI": uses_ai,
            "RiskLevel": risk_level,
            "ReviewStatus": "NotReviewed",
            "AICapabilities": None,
            "ReviewDate": None,
        }
        
        # Col 33 - Used in capability mapping
        cap_mapped = safe_str(ws.cell(row=row_idx, column=33).value)
        if cap_mapped and cap_mapped.lower() == "no":
            app["OrganizationTags"].append("NotCapabilityMapped")
        
        # Col 34 - Platform → PlatformTags
        platform = safe_str(ws.cell(row=row_idx, column=34).value)
        app["PlatformTags"] = [p.strip() for p in platform.split(",")] if platform else []
        
        # Cols 35-38 - Clean up fields → CustomFields
        custom_fields = {}
        cleanup_note = safe_str(ws.cell(row=row_idx, column=35).value)
        if cleanup_note:
            custom_fields["CleanUp_Note"] = cleanup_note
        
        cleanup_rec = safe_str(ws.cell(row=row_idx, column=36).value)
        if cleanup_rec:
            custom_fields["CleanUp_Recommendation"] = cleanup_rec
        
        cleanup_reason = safe_str(ws.cell(row=row_idx, column=37).value)
        if cleanup_reason:
            custom_fields["CleanUp_Reason"] = cleanup_reason
        
        cleanup_wp = safe_str(ws.cell(row=row_idx, column=38).value)
        if cleanup_wp:
            custom_fields["CleanUp_WPSize"] = cleanup_wp
        
        # Col 11 - ItemNumber → custom field
        item_number = safe_str(ws.cell(row=row_idx, column=11).value)
        if item_number:
            custom_fields["ItemNumber"] = item_number
        
        # Col 24 - Named users → custom field
        named_users = safe_str(ws.cell(row=row_idx, column=24).value)
        if named_users:
            custom_fields["NamedUsers"] = named_users
        
        # Col 28 - License responsible → custom field
        lic_responsible = safe_str(ws.cell(row=row_idx, column=28).value)
        if lic_responsible:
            custom_fields["LicenseResponsible"] = lic_responsible
        
        app["_CustomFields"] = custom_fields
        
        # Cols 39-41 - Security Assessment
        sec_approved_raw = safe_str(ws.cell(row=row_idx, column=39).value)
        sec_status = "NotAssessed"
        if sec_approved_raw and sec_approved_raw in SECURITY_STATUS_MAPPING:
            sec_status = SECURITY_STATUS_MAPPING[sec_approved_raw]
            report.count(f"security_{sec_status}")
        
        sec_date = safe_date(ws.cell(row=row_idx, column=40).value)
        
        sec_debt_raw = safe_str(ws.cell(row=row_idx, column=41).value)
        sec_debt_enum = "Unknown"
        if sec_debt_raw and sec_debt_raw in SECURITY_DEBT_MAPPING:
            sec_debt_enum = SECURITY_DEBT_MAPPING[sec_debt_raw]
            report.count(f"secdebt_{sec_debt_enum}")
        
        app["SecurityAssessment"] = {
            "Status": sec_status,
            "ApprovedDate": sec_date,
            "NextReviewDate": None,
            "DebtLevel": sec_debt_enum,
            "Notes": None,
        }
        
        # CreationMetadata
        app["CreationMetadata"] = {
            "Source": "CSVImport",
            "CreatedAt": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "CreatedBy": "KamstrupImport",
            "SourceIdentifier": f"Kamstrup A-Applications Row {row_idx}, NoApp={no_application}",
            "AdditionalMetadata": {
                "ImportPhase": "Phase3",
                "SourceSheet": "A-Applications",
                "RowNumber": str(row_idx),
                "NoApplication": no_application or "",
            }
        }
        
        app["Retired"] = False
        app["IsBusinessApplication"] = True
        
        apps.append(app)
    
    report.count("active_total", row_count)
    report.log(f"  Parsed {row_count} active applications")
    
    # ── Second pass: Resolve hierarchy (Sub component to) ──
    report.log(f"\n── Resolving Hierarchy ──")
    hierarchy_resolved = 0
    hierarchy_unresolved = 0
    
    # Build lookup: NoApplication string → GUID
    noapp_to_guid = {}
    for a in apps:
        if a["_NoApplication"]:
            noapp_to_guid[a["_NoApplication"]] = a["Id"]
        noapp_to_guid[a["DisplayName"]] = a["Id"]
    
    for child_guid, parent_ref, child_name in sub_component_refs:
        parent_guid = noapp_to_guid.get(parent_ref)
        if parent_guid:
            # Skip self-references (app is sub-component of itself)
            if parent_guid == child_guid:
                report.warn(f"Self-reference skipped: '{child_name}' references itself via '{parent_ref}'")
                continue
            # Set parent reference on child
            child_app = next((a for a in apps if a["Id"] == child_guid), None)
            if child_app:
                child_app["_HierarchyParentId"] = parent_guid
                # Also update parent's HierarchyChildrenIds
                parent_app = next((a for a in apps if a["Id"] == parent_guid), None)
                if parent_app:
                    if "_HierarchyChildrenIds" not in parent_app:
                        parent_app["_HierarchyChildrenIds"] = []
                    parent_app["_HierarchyChildrenIds"].append(child_guid)
                hierarchy_resolved += 1
                report.log(f"  ✓ '{child_name}' -> parent '{parent_ref}'")
        else:
            hierarchy_unresolved += 1
            report.warn(f"Unresolved parent ref '{parent_ref}' for app '{child_name}'")
    
    report.count("hierarchy_resolved", hierarchy_resolved)
    report.count("hierarchy_unresolved", hierarchy_unresolved)
    
    # ── Second pass: Resolve successors ──
    report.log(f"\n── Resolving Successors ──")
    succ_resolved = 0
    succ_unresolved = 0
    
    for app_guid, replacement_name in successor_refs:
        repl_guid = app_name_to_guid.get(replacement_name)
        if repl_guid:
            app_obj = next((a for a in apps if a["Id"] == app_guid), None)
            if app_obj:
                app_obj["SuccessorId"].append(repl_guid)
                succ_resolved += 1
                report.log(f"  OK '{app_obj['DisplayName']}' successor -> '{replacement_name}'")
        else:
            succ_unresolved += 1
            report.warn(f"Unresolved successor '{replacement_name}' for app GUID {app_guid}")
    
    report.count("successor_resolved", succ_resolved)
    report.count("successor_unresolved", succ_unresolved)
    
    return apps, app_number_to_guid, app_name_to_guid, providers_set


# ─── Removed Applications Parser ──────────────────────────────────────────

def parse_removed_applications(ws, report: ImportReport, active_name_to_guid: dict):
    """Parse A-Applications_Removed sheet (~88 rows with data)."""
    
    apps = []
    providers_set = set()
    
    report.log(f"\n── Parsing A-Applications_Removed ──")
    report.log(f"  Sheet range: rows 4 to {ws.max_row}, {ws.max_column} columns")
    
    row_count = 0
    for row_idx in range(4, ws.max_row + 1):
        app_name = safe_str(ws.cell(row=row_idx, column=4).value)
        if not app_name:
            continue
        
        row_count += 1
        
        no_application = safe_str(ws.cell(row=row_idx, column=2).value)
        app_no = safe_int(ws.cell(row=row_idx, column=3).value)
        
        # Generate deterministic GUID (different namespace to avoid collision with active)
        if app_no is not None:
            app_guid = generate_deterministic_guid(app_no + 100000)  # Offset to differentiate
        else:
            app_guid = str(uuid.uuid4())
            report.warn(f"Removed row {row_idx}: No app number for '{app_name}', using random GUID")
        
        app = OrderedDict()
        app["Id"] = app_guid
        app["DisplayName"] = app_name
        app["_RowIndex"] = row_idx
        app["_NoApplication"] = no_application
        app["_AppNo"] = app_no
        
        # Col 5 - Application Group → Category
        app["Category"] = safe_str(ws.cell(row=row_idx, column=5).value) or ""
        
        # Col 6 - Business Area
        business_area = safe_str(ws.cell(row=row_idx, column=6).value)
        app["OrganizationTags"] = [business_area] if business_area else []
        
        # Col 8 - Date of removal → LifeCycle.EndOfLife
        removal_date = safe_date(ws.cell(row=row_idx, column=8).value)
        app["LifeCycle"] = {
            "Plan": None,
            "PhaseIn": None,
            "Active": None,
            "PhaseOut": None,
            "EndOfLife": removal_date,
        }
        app["LifecycleStage"] = "Removed"
        
        # Col 9 - Comment → RichDescription
        comment = safe_str(ws.cell(row=row_idx, column=9).value)
        app["RichDescription"] = comment
        
        # Col 11 - The application to take over → SuccessorId
        successor_name = safe_str(ws.cell(row=row_idx, column=11).value)
        successor_guid = active_name_to_guid.get(successor_name) if successor_name else None
        app["SuccessorId"] = [successor_guid] if successor_guid else []
        if successor_name and not successor_guid:
            report.warn(f"Removed row {row_idx}: Unresolved successor '{successor_name}' for '{app_name}'")
        
        # Col 15 - Supplier → Provider
        supplier = safe_str(ws.cell(row=row_idx, column=15).value)
        if supplier:
            providers_set.add(supplier)
        
        # Col 27 - Vendor → Provider
        vendor = safe_str(ws.cell(row=row_idx, column=27).value)
        if vendor:
            providers_set.add(vendor)
        
        app["_Supplier"] = supplier
        app["_Vendor"] = vendor
        
        # Custom fields for financial data
        custom_fields = {}
        
        yearly_saving = ws.cell(row=row_idx, column=7).value
        if yearly_saving is not None:
            custom_fields["YearlySavingDKR"] = str(yearly_saving)
        
        yearly_cost_takeover = ws.cell(row=row_idx, column=12).value
        if yearly_cost_takeover is not None:
            custom_fields["YearlyCostTakeover"] = str(yearly_cost_takeover)
        
        cost_type = safe_str(ws.cell(row=row_idx, column=17).value)
        if cost_type:
            custom_fields["CostType"] = cost_type
        
        currency = safe_str(ws.cell(row=row_idx, column=24).value)
        if currency:
            custom_fields["Currency"] = currency
        
        yearly_costs = ws.cell(row=row_idx, column=25).value
        if yearly_costs is not None:
            custom_fields["YearlyCosts"] = str(yearly_costs)
        
        country = safe_str(ws.cell(row=row_idx, column=34).value)
        if country:
            custom_fields["Country"] = country
        
        termination_date = safe_date(ws.cell(row=row_idx, column=35).value)
        if termination_date:
            custom_fields["TerminationDate"] = termination_date
        
        actual_saving = ws.cell(row=row_idx, column=38).value
        if actual_saving is not None:
            custom_fields["ActualSavingAnnual"] = str(actual_saving)
        
        validation = safe_str(ws.cell(row=row_idx, column=39).value)
        if validation:
            custom_fields["ValidationStatus"] = validation
        
        # Cap mapping
        cap_mapped = safe_str(ws.cell(row=row_idx, column=30).value)
        if cap_mapped:
            custom_fields["UsedInCapabilityMapping"] = cap_mapped
        
        app["_CustomFields"] = custom_fields
        
        # Mark as retired
        app["Retired"] = True
        
        # Default empty fields
        app["HostedOn"] = {"Criticality": "Unmapped", "HostingTypeValue": None, "HostingTypeDescription": None}
        app["PortFolioStrategy"] = {"TIME": "Unmapped", "Description": None}
        app["ApplicationTypeTags"] = []
        app["UserBase"] = {"Size": "Unknown", "Count": None, "Description": None, "UserType": "Unknown"}
        app["AIClassification"] = {"UsesAI": False, "RiskLevel": "NotApplicable", "ReviewStatus": "NotReviewed", "AICapabilities": None, "ReviewDate": None}
        app["SecurityAssessment"] = {"Status": "NotAssessed", "ApprovedDate": None, "NextReviewDate": None, "DebtLevel": "Unknown", "Notes": None}
        app["PlatformTags"] = []
        app["Owners"] = []
        app["Responsible"] = None
        app["Url"] = None
        
        app["CreationMetadata"] = {
            "Source": "CSVImport",
            "CreatedAt": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "CreatedBy": "KamstrupImport",
            "SourceIdentifier": f"Kamstrup A-Applications_Removed Row {row_idx}, NoApp={no_application}",
            "AdditionalMetadata": {
                "ImportPhase": "Phase3",
                "SourceSheet": "A-Applications_Removed",
                "RowNumber": str(row_idx),
                "NoApplication": no_application or "",
            }
        }
        
        apps.append(app)
    
    report.count("removed_total", row_count)
    report.log(f"  Parsed {row_count} removed applications")
    
    return apps, providers_set


# ─── Provider Extraction ──────────────────────────────────────────────────

def build_provider_factsheets(providers_set: set, report: ImportReport):
    """Create ProviderFactSheet records from unique vendor/consultancy names."""
    
    PROVIDER_NAMESPACE = uuid.UUID("b2c3d4e5-f6a7-8901-bcde-f01234567890")
    
    providers = []
    for name in sorted(providers_set):
        if not name:
            continue
        provider_guid = str(uuid.uuid5(PROVIDER_NAMESPACE, name))
        providers.append({
            "Id": provider_guid,
            "DisplayName": name,
            "ProviderType": "Software",  # Default; can be refined
            "CreationMetadata": {
                "Source": "CSVImport",
                "CreatedAt": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
                "CreatedBy": "KamstrupImport",
                "SourceIdentifier": f"Extracted from A-Applications vendor/consultancy columns",
            }
        })
    
    report.count("providers_total", len(providers))
    report.log(f"\n── Provider Extraction ──")
    report.log(f"  Extracted {len(providers)} unique providers")
    
    return providers


# ─── Custom Column Definitions ─────────────────────────────────────────────

def build_custom_column_definitions():
    """Generate CustomColumnDefinition records for Kamstrup-specific fields."""
    
    definitions = [
        {
            "Id": str(uuid.uuid5(KAMSTRUP_APP_NAMESPACE, "CleanUp_Note")),
            "Name": "Clean up: Note",
            "Description": "Kamstrup cleanup tracking note",
            "FactSheetType": "Application",
            "FieldType": "TextArea",
            "Required": False,
            "DisplayOrder": 200,
            "VisibleByDefault": False,
        },
        {
            "Id": str(uuid.uuid5(KAMSTRUP_APP_NAMESPACE, "CleanUp_Recommendation")),
            "Name": "Clean up: Recommendation",
            "Description": "Kamstrup cleanup recommendation",
            "FactSheetType": "Application",
            "FieldType": "TextArea",
            "Required": False,
            "DisplayOrder": 201,
            "VisibleByDefault": False,
        },
        {
            "Id": str(uuid.uuid5(KAMSTRUP_APP_NAMESPACE, "CleanUp_Reason")),
            "Name": "Clean up: Reason",
            "Description": "Kamstrup cleanup reason",
            "FactSheetType": "Application",
            "FieldType": "TextArea",
            "Required": False,
            "DisplayOrder": 202,
            "VisibleByDefault": False,
        },
        {
            "Id": str(uuid.uuid5(KAMSTRUP_APP_NAMESPACE, "CleanUp_WPSize")),
            "Name": "Clean up: WP Size",
            "Description": "Work package size for cleanup",
            "FactSheetType": "Application",
            "FieldType": "Dropdown",
            "Required": False,
            "AllowedValues": ["S", "M", "L", "Project", "Project-initiated"],
            "DisplayOrder": 203,
            "VisibleByDefault": False,
        },
        {
            "Id": str(uuid.uuid5(KAMSTRUP_APP_NAMESPACE, "ItemNumber")),
            "Name": "Item Number",
            "Description": "Kamstrup internal item number",
            "FactSheetType": "Application",
            "FieldType": "Text",
            "Required": False,
            "DisplayOrder": 210,
            "VisibleByDefault": False,
        },
        {
            "Id": str(uuid.uuid5(KAMSTRUP_APP_NAMESPACE, "NamedUsers")),
            "Name": "Named Users",
            "Description": "Named users of the application",
            "FactSheetType": "Application",
            "FieldType": "TextArea",
            "Required": False,
            "DisplayOrder": 211,
            "VisibleByDefault": False,
        },
        {
            "Id": str(uuid.uuid5(KAMSTRUP_APP_NAMESPACE, "LicenseResponsible")),
            "Name": "License Responsible",
            "Description": "Person responsible for license management",
            "FactSheetType": "Application",
            "FieldType": "Text",
            "Required": False,
            "DisplayOrder": 212,
            "VisibleByDefault": False,
        },
        # Financial fields for removed apps
        {
            "Id": str(uuid.uuid5(KAMSTRUP_APP_NAMESPACE, "YearlySavingDKR")),
            "Name": "Yearly Saving (DKR)",
            "Description": "Annual cost saving from application removal",
            "FactSheetType": "Application",
            "FieldType": "Currency",
            "Required": False,
            "DisplayOrder": 220,
            "VisibleByDefault": False,
        },
        {
            "Id": str(uuid.uuid5(KAMSTRUP_APP_NAMESPACE, "YearlyCosts")),
            "Name": "Yearly Costs",
            "Description": "Annual cost of the application",
            "FactSheetType": "Application",
            "FieldType": "Currency",
            "Required": False,
            "DisplayOrder": 221,
            "VisibleByDefault": False,
        },
        {
            "Id": str(uuid.uuid5(KAMSTRUP_APP_NAMESPACE, "ActualSavingAnnual")),
            "Name": "Actual Saving (Annual)",
            "Description": "Actual annual saving from removal",
            "FactSheetType": "Application",
            "FieldType": "Currency",
            "Required": False,
            "DisplayOrder": 222,
            "VisibleByDefault": False,
        },
    ]
    
    return definitions


# ─── Output Transformer ───────────────────────────────────────────────────

def transform_for_omnigaze_api(app: dict, custom_col_defs: list) -> dict:
    """Transform internal representation to OmniGaze-compatible JSON.
    
    This produces a structure that maps directly to the C# model properties,
    suitable for use with the MCP API CreateFactSheet tool or direct
    FactSheetContainer.AddIfNotExists() via a C# loader.
    """
    
    # Build custom field values list using deterministic column IDs
    custom_fields = []
    custom_field_data = app.get("_CustomFields", {})
    col_name_to_id = {d["Name"].replace("Clean up: ", "CleanUp_").replace(" ", "").replace("(", "").replace(")", "").replace(",", ""): d["Id"] for d in custom_col_defs}
    
    # Build a simpler lookup
    CUSTOM_FIELD_KEY_MAP = {}
    for d in custom_col_defs:
        # Create a map from our internal key names to definition IDs
        key = d["Name"]
        CUSTOM_FIELD_KEY_MAP[key] = d["Id"]
    
    # Map internal keys to column definition names
    INTERNAL_TO_DEF = {
        "CleanUp_Note": "Clean up: Note",
        "CleanUp_Recommendation": "Clean up: Recommendation",
        "CleanUp_Reason": "Clean up: Reason",
        "CleanUp_WPSize": "Clean up: WP Size",
        "ItemNumber": "Item Number",
        "NamedUsers": "Named Users",
        "LicenseResponsible": "License Responsible",
        "YearlySavingDKR": "Yearly Saving (DKR)",
        "YearlyCosts": "Yearly Costs",
        "ActualSavingAnnual": "Actual Saving (Annual)",
        # Pass-through for others not in definitions
        "CostType": None,
        "Currency": None,
        "Country": None,
        "TerminationDate": None,
        "ValidationStatus": None,
        "UsedInCapabilityMapping": None,
        "YearlyCostTakeover": None,
    }
    
    for field_key, field_value in custom_field_data.items():
        def_name = INTERNAL_TO_DEF.get(field_key)
        if def_name and def_name in CUSTOM_FIELD_KEY_MAP:
            custom_fields.append({
                "ColumnId": CUSTOM_FIELD_KEY_MAP[def_name],
                "Value": field_value,
            })
        elif def_name is None:
            # Field not in definitions, store in AdditionalMetadata instead
            if "CreationMetadata" in app and "AdditionalMetadata" in app["CreationMetadata"]:
                app["CreationMetadata"]["AdditionalMetadata"][field_key] = str(field_value)
    
    # Build hierarchy IDs
    hierarchy_children_ids = app.get("_HierarchyChildrenIds", [])
    
    result = OrderedDict()
    result["Id"] = app["Id"]
    result["DisplayName"] = app["DisplayName"]
    result["Category"] = app.get("Category", "")
    result["RichDescription"] = app.get("RichDescription")
    result["Url"] = app.get("Url")
    result["Responsible"] = app.get("Responsible")
    result["Retired"] = app.get("Retired", False)
    result["LifeCycle"] = app["LifeCycle"]
    result["LifecycleStage"] = app.get("LifecycleStage")
    result["PortFolioStrategy"] = app["PortFolioStrategy"]
    result["HostedOn"] = app["HostedOn"]
    result["ApplicationTypeTags"] = app.get("ApplicationTypeTags", [])
    result["PlatformTags"] = app.get("PlatformTags", [])
    result["OrganizationTags"] = app.get("OrganizationTags", [])
    result["UserBase"] = app["UserBase"]
    result["AIClassification"] = app["AIClassification"]
    result["SecurityAssessment"] = app["SecurityAssessment"]
    result["Owners"] = app.get("Owners", [])
    result["SuccessorId"] = app.get("SuccessorId", [])
    result["CustomFields"] = custom_fields
    result["CreationMetadata"] = app.get("CreationMetadata")
    result["HierarchyChildrenIds"] = hierarchy_children_ids
    
    # Pass-through metadata for Phase 4
    result["_SoftwareVendor"] = app.get("_SoftwareVendor")
    result["_Consultancy"] = app.get("_Consultancy")
    result["_SupportingApplications"] = app.get("_SupportingApplications")
    result["_Org"] = app.get("_Org")
    result["_SuperUser"] = app.get("_SuperUser")
    result["_NoApplication"] = app.get("_NoApplication")
    result["_AppNo"] = app.get("_AppNo")
    result["_HierarchyParentId"] = app.get("_HierarchyParentId")
    
    return result


# ─── Main ─────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Kamstrup Phase 3: Application Import")
    parser.add_argument("--excel", default=EXCEL_PATH, help="Path to Kamstrup Excel file")
    parser.add_argument("--output-dir", default=OUTPUT_DIR, help="Output directory")
    args = parser.parse_args()
    
    os.makedirs(args.output_dir, exist_ok=True)
    
    report = ImportReport()
    report.log(f"Excel source: {args.excel}")
    report.log(f"Output directory: {args.output_dir}")
    report.log(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Load workbook
    print(f"Loading workbook: {args.excel}")
    wb = openpyxl.load_workbook(args.excel, data_only=True)
    
    # Parse active applications
    print("Parsing A-Applications...")
    ws_active = wb["A-Applications"]
    active_apps, app_number_to_guid, app_name_to_guid, active_providers = parse_active_applications(ws_active, report)
    
    # Parse removed applications
    print("Parsing A-Applications_Removed...")
    ws_removed = wb["A-Applications_Removed"]
    removed_apps, removed_providers = parse_removed_applications(ws_removed, report, app_name_to_guid)
    
    # Merge providers
    all_providers = active_providers | removed_providers
    providers = build_provider_factsheets(all_providers, report)
    
    # Custom column definitions
    custom_col_defs = build_custom_column_definitions()
    
    # Transform to OmniGaze format
    print("Transforming to OmniGaze format...")
    active_output = [transform_for_omnigaze_api(a, custom_col_defs) for a in active_apps]
    removed_output = [transform_for_omnigaze_api(a, custom_col_defs) for a in removed_apps]
    
    # Summary stats
    report.log(f"\n── Summary ──")
    report.log(f"  Active applications: {len(active_output)}")
    report.log(f"  Removed applications: {len(removed_output)}")
    report.log(f"  Total applications: {len(active_output) + len(removed_output)}")
    report.log(f"  Unique providers: {len(providers)}")
    report.log(f"  Custom column definitions: {len(custom_col_defs)}")
    
    # Lifecycle distribution
    lc_dist = Counter()
    for a in active_output:
        ls = a.get("LifecycleStage")
        if ls:
            lc_dist[ls] += 1
    report.log(f"\n── Lifecycle Distribution ──")
    for k, v in sorted(lc_dist.items()):
        report.log(f"  {k}: {v}")
    
    # Write output files
    print("Writing output files...")
    
    def write_json(filename, data):
        path = os.path.join(args.output_dir, filename)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)
        size_kb = os.path.getsize(path) / 1024
        print(f"  -> {filename} ({size_kb:.1f} KB)")
        report.log(f"  Wrote {filename} ({size_kb:.1f} KB)")
    
    write_json("applications_active.json", active_output)
    write_json("applications_removed.json", removed_output)
    write_json("providers_extracted.json", providers)
    write_json("app_number_to_guid.json", app_number_to_guid)
    write_json("custom_column_definitions.json", custom_col_defs)
    
    # Write report
    report_text = report.dump()
    report_path = os.path.join(args.output_dir, "import_report.txt")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report_text)
    print(f"  -> import_report.txt")
    
    # Print summary
    print("\n" + "=" * 60)
    print("  Phase 3: Application Import Complete")
    print("=" * 60)
    print(f"  Active:   {len(active_output)} applications")
    print(f"  Removed:  {len(removed_output)} applications")
    print(f"  Total:    {len(active_output) + len(removed_output)} applications")
    print(f"  Providers: {len(providers)}")
    print(f"  Warnings: {len(report.warnings)}")
    print(f"  Errors:   {len(report.errors)}")
    print("=" * 60)
    
    if report.errors:
        print("\n❌ ERRORS:")
        for e in report.errors:
            print(f"  {e}")
    
    return 0 if not report.errors else 1


if __name__ == "__main__":
    sys.exit(main())
