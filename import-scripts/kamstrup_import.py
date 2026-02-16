#!/usr/bin/env python3
"""
Kamstrup Master Data Import — Phase 2
======================================
Parses the Kamstrup_Business-Capability-Map.xlsx file and generates
structured JSON files ready for import into OmniGaze's FactSheetContainer.

Each JSON file contains factsheet records that map 1:1 to OmniGaze model classes.
A deterministic GUID is generated for each entity from its composite key so that
Phase 4 (relationships) can wire things together without re-parsing.

Usage:
    python kamstrup_import.py [--excel PATH] [--output DIR]
"""

import argparse
import hashlib
import json
import os
import sys
import uuid
from collections import OrderedDict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# ─── Configuration ─────────────────────────────────────────────────────────────

DEFAULT_EXCEL = r"F:\RootContext\Kamstrup\Kamstrup_Business-Capability-Map.xlsx"
DEFAULT_OUTPUT = r"F:\RootContext\Kamstrup\import-scripts\output"

# Deterministic GUID namespace (fixed UUID5 namespace for Kamstrup import)
NAMESPACE_KAMSTRUP = uuid.UUID("a1b2c3d4-e5f6-7890-abcd-ef1234567890")

# ─── Helpers ───────────────────────────────────────────────────────────────────

def deterministic_guid(category: str, key: str) -> str:
    """Generate a deterministic GUID from category + key using UUID5."""
    composite = f"{category}::{key}"
    return str(uuid.uuid5(NAMESPACE_KAMSTRUP, composite))


def safe_str(value) -> str:
    """Convert cell value to trimmed string, or empty string if None."""
    if value is None:
        return ""
    return str(value).strip()


def safe_str_or_none(value):
    """Return trimmed string if non-empty, else None."""
    s = safe_str(value)
    return s if s else None


def map_pace_layer(value: str) -> int:
    """Map Kamstrup pace layer string to PaceLayerEnum int."""
    mapping = {
        "systems of innovation": 1,
        "systems of differentiation": 2,
        "systems of commodity": 3,
    }
    return mapping.get(safe_str(value).lower(), 0)


def map_nis2_criticality(value: str) -> int:
    """Map NIS2 criticality string to NIS2CriticalityEnum int."""
    mapping = {
        "high": 3,
        "medium": 2,
        "low": 1,
        "not evaluated": 0,
    }
    return mapping.get(safe_str(value).lower(), 0)


def map_editing_state(value: str) -> str:
    """Return the editing state as a string (stored as configurable list value)."""
    val = safe_str(value)
    if not val or val.lower() == "n/a":
        return "N/A"
    return val


class ImportLog:
    """Collects import statistics and warnings."""
    def __init__(self):
        self.entries = []
        self.warnings = []
        self.counts = {}

    def info(self, msg):
        self.entries.append(f"[INFO] {msg}")
        print(f"  [INFO] {msg}")

    def warn(self, msg):
        self.warnings.append(msg)
        self.entries.append(f"[WARN] {msg}")
        print(f"  [WARN] {msg}")

    def count(self, category, n):
        self.counts[category] = n
        self.entries.append(f"[COUNT] {category}: {n}")
        print(f"  [COUNT] {category}: {n}")

    def save(self, path):
        with open(path, "w", encoding="utf-8") as f:
            f.write(f"Kamstrup Phase 2 Import Log\n")
            f.write(f"Generated: {datetime.now().isoformat()}\n")
            f.write(f"{'='*60}\n\n")
            f.write("SUMMARY\n")
            f.write("-" * 40 + "\n")
            for k, v in self.counts.items():
                f.write(f"  {k}: {v}\n")
            total = sum(self.counts.values())
            f.write(f"  TOTAL: {total}\n\n")
            if self.warnings:
                f.write(f"WARNINGS ({len(self.warnings)})\n")
                f.write("-" * 40 + "\n")
                for w in self.warnings:
                    f.write(f"  ⚠ {w}\n")
                f.write("\n")
            f.write("DETAILED LOG\n")
            f.write("-" * 40 + "\n")
            for e in self.entries:
                f.write(f"  {e}\n")


log = ImportLog()


# ─── Import Functions ──────────────────────────────────────────────────────────

def import_value_streams(wb) -> dict:
    """
    Extract the 12 unique value streams from B-Business Capability (col K/11)
    and C-Business Context (col P/16).
    Returns: dict of display_name -> factsheet dict
    """
    print("\n[1/6] Importing Value Streams...")
    vs_names = set()

    # From B-Business Capability col 11
    ws = wb["B-Business Capability"]
    for row in ws.iter_rows(min_row=6, values_only=True):
        val = safe_str(row[10] if len(row) > 10 else None)  # col K = index 10
        if val and val.lower() != "n/a":
            vs_names.add(val)

    # From C-Business Context col 16
    ws2 = wb["C-Business Context"]
    for row in ws2.iter_rows(min_row=3, values_only=True):
        val = safe_str(row[15] if len(row) > 15 else None)  # col P = index 15
        if val and val.lower() != "n/a":
            vs_names.add(val)

    result = {}
    for name in sorted(vs_names):
        guid = deterministic_guid("ValueStream", name)
        result[name] = {
            "Id": guid,
            "DisplayName": name,
            "FactSheetType": "ValueStreamFactSheet",
        }

    log.count("ValueStreams", len(result))
    log.info(f"Value Streams: {sorted(vs_names)}")
    return result


def import_business_capabilities(wb, vs_lookup: dict) -> tuple:
    """
    Import B-Business Capability sheet.
    Returns: (list_of_factsheets, id_lookup_dict)
    """
    print("\n[2/6] Importing Business Capabilities...")
    ws = wb["B-Business Capability"]

    # Collect all rows first
    rows = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        id_key = safe_str(row[1] if len(row) > 1 else None)  # col B = index 1
        if not id_key:
            continue
        rows.append({
            "id_key": id_key,
            "type": safe_str(row[3] if len(row) > 3 else None),
            "level0": safe_str(row[4] if len(row) > 4 else None),
            "level1": safe_str(row[5] if len(row) > 5 else None),
            "level2": safe_str(row[6] if len(row) > 6 else None),
            "level3": safe_str(row[7] if len(row) > 7 else None),
            "pace_layer": safe_str(row[8] if len(row) > 8 else None),
            "description": safe_str(row[9] if len(row) > 9 else None),
            "value_stream": safe_str(row[10] if len(row) > 10 else None),
            "vs_order": safe_str(row[11] if len(row) > 11 else None),
            "nis2": safe_str(row[12] if len(row) > 12 else None),
            "updated": safe_str(row[13] if len(row) > 13 else None),
            "responsible": safe_str(row[14] if len(row) > 14 else None),
        })

    log.info(f"Read {len(rows)} data rows from B-Business Capability")

    # Pass 0: L0 groups (col E/5) — import as top-level grouping capabilities
    l0_names = sorted(set(r["level0"] for r in rows if r["level0"]))
    l0_lookup = {}  # name -> guid
    factsheets = []

    for name in l0_names:
        guid = deterministic_guid("BC_L0", name)
        l0_lookup[name] = guid
        factsheets.append({
            "Id": guid,
            "DisplayName": name,
            "FactSheetType": "BusinessCapabilityFactSheet",
            "HierarchyLevel": 0,
            "HierarchyParentId": None,
            "PaceLayer": 0,
            "NIS2Criticality": 0,
        })

    log.info(f"L0 groups: {len(l0_names)} — {l0_names}")

    # Pass 1: L1 capabilities (col F/6) — deduplicated, parented to L0
    l1_seen = {}  # (l0, l1) -> guid
    for r in rows:
        l1 = r["level1"]
        if not l1:
            continue
        l0 = r["level0"]
        key = (l0, l1)
        if key not in l1_seen:
            guid = deterministic_guid("BC_L1", f"{l0}::{l1}")
            l1_seen[key] = guid
            parent_id = l0_lookup.get(l0)
            factsheets.append({
                "Id": guid,
                "DisplayName": l1,
                "FactSheetType": "BusinessCapabilityFactSheet",
                "HierarchyLevel": 1,
                "HierarchyParentId": parent_id,
                "PaceLayer": map_pace_layer(r["pace_layer"]) if not r["level2"] else 0,
                "NIS2Criticality": map_nis2_criticality(r["nis2"]) if not r["level2"] else 0,
                "RichDescription": r["description"] if not r["level2"] and r["description"] else None,
                "Responsible": r["responsible"] if not r["level2"] and r["responsible"] else None,
            })

    log.info(f"L1 capabilities: {len(l1_seen)}")

    # Pass 2: L2 capabilities (col G/7)
    l2_seen = {}  # (l0, l1, l2) -> guid
    for r in rows:
        l2 = r["level2"]
        if not l2:
            continue
        l0, l1 = r["level0"], r["level1"]
        key = (l0, l1, l2)
        if key not in l2_seen:
            guid = deterministic_guid("BC_L2", f"{l0}::{l1}::{l2}")
            l2_seen[key] = guid
            parent_id = l1_seen.get((l0, l1))
            factsheets.append({
                "Id": guid,
                "DisplayName": l2,
                "FactSheetType": "BusinessCapabilityFactSheet",
                "HierarchyLevel": 2,
                "HierarchyParentId": parent_id,
                "PaceLayer": map_pace_layer(r["pace_layer"]) if not r["level3"] else 0,
                "NIS2Criticality": map_nis2_criticality(r["nis2"]) if not r["level3"] else 0,
                "RichDescription": r["description"] if not r["level3"] and r["description"] else None,
                "Responsible": r["responsible"] if not r["level3"] and r["responsible"] else None,
            })

    log.info(f"L2 capabilities: {len(l2_seen)}")

    # Pass 3: L3 capabilities (col H/8)
    l3_seen = {}  # (l0, l1, l2, l3) -> guid
    for r in rows:
        l3 = r["level3"]
        if not l3:
            continue
        l0, l1, l2 = r["level0"], r["level1"], r["level2"]
        key = (l0, l1, l2, l3)
        if key not in l3_seen:
            guid = deterministic_guid("BC_L3", f"{l0}::{l1}::{l2}::{l3}")
            l3_seen[key] = guid
            parent_id = l2_seen.get((l0, l1, l2))
            if not parent_id:
                parent_id = l1_seen.get((l0, l1))
                if parent_id:
                    log.warn(f"L3 '{l3}' has no L2 parent, falling back to L1 '{l1}'")
            
            # Find the row data for this L3 — use the current row's attributes
            factsheets.append({
                "Id": guid,
                "DisplayName": l3,
                "FactSheetType": "BusinessCapabilityFactSheet",
                "HierarchyLevel": 3,
                "HierarchyParentId": parent_id,
                "PaceLayer": map_pace_layer(r["pace_layer"]),
                "NIS2Criticality": map_nis2_criticality(r["nis2"]),
                "RichDescription": r["description"] if r["description"] else None,
                "Responsible": r["responsible"] if r["responsible"] else None,
            })

    log.info(f"L3 capabilities: {len(l3_seen)}")

    # Build value stream references on L1/L2/L3 capabilities
    # Each row maps id_key -> value_stream
    vs_refs = {}  # guid -> list of value stream guids
    for r in rows:
        vs_name = r["value_stream"]
        if not vs_name or vs_name.lower() == "n/a":
            continue
        vs_guid = vs_lookup.get(vs_name, {}).get("Id")
        if not vs_guid:
            continue
        # Find which level this row creates
        if r["level3"]:
            bc_guid = l3_seen.get((r["level0"], r["level1"], r["level2"], r["level3"]))
        elif r["level2"]:
            bc_guid = l2_seen.get((r["level0"], r["level1"], r["level2"]))
        else:
            bc_guid = l1_seen.get((r["level0"], r["level1"]))
        if bc_guid:
            if bc_guid not in vs_refs:
                vs_refs[bc_guid] = []
            if vs_guid not in vs_refs[bc_guid]:
                vs_refs[bc_guid].append(vs_guid)

    # Attach value stream refs
    for fs in factsheets:
        if fs["Id"] in vs_refs:
            fs["SupportedValueStreamIds"] = vs_refs[fs["Id"]]

    # Build composite key -> GUID lookup for Phase 4
    id_lookup = {}
    for r in rows:
        id_key = r["id_key"]
        if r["level3"]:
            guid = l3_seen.get((r["level0"], r["level1"], r["level2"], r["level3"]))
        elif r["level2"]:
            guid = l2_seen.get((r["level0"], r["level1"], r["level2"]))
        else:
            guid = l1_seen.get((r["level0"], r["level1"]))
        if guid:
            id_lookup[id_key] = guid

    total = len(l0_names) + len(l1_seen) + len(l2_seen) + len(l3_seen)
    log.count("BusinessCapabilities", total)
    log.info(f"Total BC factsheets: {total} (L0={len(l0_names)}, L1={len(l1_seen)}, L2={len(l2_seen)}, L3={len(l3_seen)})")

    return factsheets, id_lookup


def import_organizations(wb) -> tuple:
    """
    Import O-Organization sheet.
    Returns: (list_of_factsheets, org_lookup_dict)
    """
    print("\n[3/6] Importing Organizations...")
    ws = wb["O-Organization"]

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        org_bu = safe_str(row[1] if len(row) > 1 else None)        # col B = Level 2 (Org/BU)
        biz_area = safe_str(row[2] if len(row) > 2 else None)       # col C = Level 3 (Business Area)
        country = safe_str(row[3] if len(row) > 3 else None)        # col D = Level 1 (Country)
        country_code = safe_str(row[8] if len(row) > 8 else None)   # col I = CountryCode
        country_name = safe_str(row[9] if len(row) > 9 else None)   # col J = Country display name
        dept_team = safe_str(row[13] if len(row) > 13 else None)    # col N = DepartmentTeam (Level 4)

        if not country and not org_bu:
            continue

        rows.append({
            "country": country,
            "country_code": country_code,
            "country_name": country_name,
            "org_bu": org_bu,
            "biz_area": biz_area,
            "dept_team": dept_team,
        })

    log.info(f"Read {len(rows)} data rows from O-Organization")

    factsheets = []

    # Pass 1: Countries (L1)
    countries = {}
    for r in rows:
        c = r["country"]
        if c and c not in countries:
            guid = deterministic_guid("Org_Country", c)
            countries[c] = guid
            # Try to get country code from the row
            cc = r["country_code"] if r["country"] == r.get("country_name", "") else ""
            # Actually map country to code from unique rows where this country appears
            factsheets.append({
                "Id": guid,
                "DisplayName": c,
                "FactSheetType": "OrganizationFactSheet",
                "OrgType": "Country",
                "CountryCode": "",
                "HierarchyParentId": None,
                "HierarchyLevel": 1,
            })

    # Enrich country codes from the data (col I maps to col J country names)
    # The country_code and country_name columns are separate lookup columns
    country_code_map = {}  # country_name -> country_code
    for r in rows:
        cn = r["country_name"]
        cc = r["country_code"]
        if cn and cc:
            country_code_map[cn] = cc
    
    # Also map from the main country column (col D) to the first available code
    for r in rows:
        c = r["country"]
        if c and c not in country_code_map:
            # Try to find in country_name map
            pass
    
    # Apply country codes
    for fs in factsheets:
        if fs["FactSheetType"] == "OrganizationFactSheet" and fs["OrgType"] == "Country":
            cc = country_code_map.get(fs["DisplayName"], "")
            fs["CountryCode"] = cc

    log.info(f"L1 Countries: {len(countries)} — {sorted(countries.keys())}")

    # Pass 2: Business Units (L2)
    bus_units = {}
    for r in rows:
        bu = r["org_bu"]
        c = r["country"]
        if not bu or not c:
            continue
        key = (c, bu)
        if key not in bus_units:
            guid = deterministic_guid("Org_BU", f"{c}::{bu}")
            bus_units[key] = guid
            factsheets.append({
                "Id": guid,
                "DisplayName": bu,
                "FactSheetType": "OrganizationFactSheet",
                "OrgType": "Business Unit",
                "CountryCode": country_code_map.get(c, ""),
                "HierarchyParentId": countries.get(c),
                "HierarchyLevel": 2,
            })

    log.info(f"L2 Business Units: {len(bus_units)}")

    # Pass 3: Business Areas / Departments (L3)
    departments = {}
    for r in rows:
        ba = r["biz_area"]
        bu = r["org_bu"]
        c = r["country"]
        if not ba or not bu or not c:
            continue
        key = (c, bu, ba)
        if key not in departments:
            guid = deterministic_guid("Org_Dept", f"{c}::{bu}::{ba}")
            departments[key] = guid
            factsheets.append({
                "Id": guid,
                "DisplayName": ba,
                "FactSheetType": "OrganizationFactSheet",
                "OrgType": "Department",
                "CountryCode": country_code_map.get(c, ""),
                "HierarchyParentId": bus_units.get((c, bu)),
                "HierarchyLevel": 3,
            })

    log.info(f"L3 Departments/Areas: {len(departments)}")

    # Pass 4: Teams (L4) from col N
    teams = {}
    for r in rows:
        team = r["dept_team"]
        ba = r["biz_area"]
        bu = r["org_bu"]
        c = r["country"]
        if not team or not ba or not bu or not c:
            continue
        key = (c, bu, ba, team)
        if key not in teams:
            guid = deterministic_guid("Org_Team", f"{c}::{bu}::{ba}::{team}")
            teams[key] = guid
            factsheets.append({
                "Id": guid,
                "DisplayName": team,
                "FactSheetType": "OrganizationFactSheet",
                "OrgType": "Team",
                "CountryCode": country_code_map.get(c, ""),
                "HierarchyParentId": departments.get((c, bu, ba)),
                "HierarchyLevel": 4,
            })

    log.info(f"L4 Teams: {len(teams)}")

    total = len(countries) + len(bus_units) + len(departments) + len(teams)
    log.count("Organizations", total)

    # Build org lookup for Phase 4
    org_lookup = {}
    for (c, bu), guid in bus_units.items():
        org_lookup[f"{c}::{bu}"] = guid
    for (c, bu, ba), guid in departments.items():
        org_lookup[f"{c}::{bu}::{ba}"] = guid
    for c, guid in countries.items():
        org_lookup[c] = guid

    return factsheets, org_lookup


def import_processes(wb, vs_lookup: dict) -> tuple:
    """
    Import C-Business Context (Processes) sheet.
    Returns: (list_of_factsheets, process_lookup_dict)
    """
    print("\n[4/6] Importing Business Context / Processes...")
    ws = wb["C-Business Context"]

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        no_process = safe_str(row[2] if len(row) > 2 else None)   # col C
        doc_no = safe_str(row[3] if len(row) > 3 else None)       # col D
        proc_group = safe_str(row[4] if len(row) > 4 else None)   # col E
        process = safe_str(row[5] if len(row) > 5 else None)      # col F
        doc_no_sub = safe_str(row[6] if len(row) > 6 else None)   # col G
        subprocess = safe_str(row[7] if len(row) > 7 else None)   # col H
        approver = safe_str(row[8] if len(row) > 8 else None)     # col I
        bu = safe_str(row[9] if len(row) > 9 else None)           # col J (for Phase 4)
        country = safe_str(row[10] if len(row) > 10 else None)    # col K (for Phase 4)
        editing_state = safe_str(row[11] if len(row) > 11 else None)  # col L
        value_stream = safe_str(row[15] if len(row) > 15 else None)   # col P
        description = safe_str(row[16] if len(row) > 16 else None)    # col Q

        if not proc_group and not process:
            continue

        rows.append({
            "no_process": no_process,
            "doc_no": doc_no,
            "proc_group": proc_group,
            "process": process,
            "doc_no_sub": doc_no_sub,
            "subprocess": subprocess,
            "approver": approver,
            "bu": bu,
            "country": country,
            "editing_state": editing_state,
            "value_stream": value_stream,
            "description": description,
        })

    log.info(f"Read {len(rows)} data rows from C-Business Context")

    factsheets = []

    # Pass 1: Process Groups (L1)
    proc_groups = {}
    for r in rows:
        pg = r["proc_group"]
        if pg and pg not in proc_groups:
            guid = deterministic_guid("Proc_L1", pg)
            proc_groups[pg] = guid
            factsheets.append({
                "Id": guid,
                "DisplayName": pg,
                "FactSheetType": "ProcessFactSheet",
                "HierarchyLevel": 1,
                "HierarchyParentId": None,
                "EditingState": None,
                "DocumentNumber": None,
                "Approver": None,
            })

    log.info(f"L1 Process Groups: {len(proc_groups)} — {sorted(proc_groups.keys())}")

    # Pass 2: Processes (L2)
    processes = {}
    for r in rows:
        p = r["process"]
        pg = r["proc_group"]
        if not p or not pg:
            continue
        key = (pg, p)
        if key not in processes:
            guid = deterministic_guid("Proc_L2", f"{pg}::{p}")
            processes[key] = guid
            factsheets.append({
                "Id": guid,
                "DisplayName": p,
                "FactSheetType": "ProcessFactSheet",
                "HierarchyLevel": 2,
                "HierarchyParentId": proc_groups.get(pg),
                "EditingState": map_editing_state(r["editing_state"]) if not r["subprocess"] else None,
                "DocumentNumber": r["doc_no"] if r["doc_no"] and not r["subprocess"] else None,
                "Approver": r["approver"] if not r["subprocess"] else None,
                "RichDescription": r["description"] if not r["subprocess"] and r["description"] else None,
            })

    log.info(f"L2 Processes: {len(processes)}")

    # Pass 3: SubProcesses (L3)
    subprocesses = {}
    for r in rows:
        sp = r["subprocess"]
        p = r["process"]
        pg = r["proc_group"]
        if not sp or not p:
            continue
        key = (pg, p, sp)
        if key not in subprocesses:
            guid = deterministic_guid("Proc_L3", f"{pg}::{p}::{sp}")
            subprocesses[key] = guid
            factsheets.append({
                "Id": guid,
                "DisplayName": sp,
                "FactSheetType": "ProcessFactSheet",
                "HierarchyLevel": 3,
                "HierarchyParentId": processes.get((pg, p)),
                "EditingState": map_editing_state(r["editing_state"]),
                "DocumentNumber": r["doc_no_sub"] if r["doc_no_sub"] else r["doc_no"],
                "Approver": r["approver"] if r["approver"] else None,
                "RichDescription": r["description"] if r["description"] else None,
            })

    log.info(f"L3 SubProcesses: {len(subprocesses)}")

    # Attach value stream references
    vs_refs = {}
    for r in rows:
        vs_name = r["value_stream"]
        if not vs_name or vs_name.lower() == "n/a":
            continue
        vs_guid = vs_lookup.get(vs_name, {}).get("Id")
        if not vs_guid:
            continue
        
        if r["subprocess"]:
            proc_guid = subprocesses.get((r["proc_group"], r["process"], r["subprocess"]))
        else:
            proc_guid = processes.get((r["proc_group"], r["process"]))
        
        if proc_guid:
            if proc_guid not in vs_refs:
                vs_refs[proc_guid] = []
            if vs_guid not in vs_refs[proc_guid]:
                vs_refs[proc_guid].append(vs_guid)

    for fs in factsheets:
        if fs["Id"] in vs_refs:
            fs["SupportedValueStreamIds"] = vs_refs[fs["Id"]]

    total = len(proc_groups) + len(processes) + len(subprocesses)
    log.count("Processes", total)

    # Build process lookup for Phase 4
    proc_lookup = {}
    for r in rows:
        np = r["no_process"]
        if not np:
            continue
        if r["subprocess"]:
            guid = subprocesses.get((r["proc_group"], r["process"], r["subprocess"]))
        else:
            guid = processes.get((r["proc_group"], r["process"]))
        if guid:
            proc_lookup[np] = guid

    return factsheets, proc_lookup


def import_platforms(wb) -> tuple:
    """
    Import P_Platform sheet → ITComponentFactSheet.
    Returns: (list_of_factsheets, platform_lookup_dict)
    """
    print("\n[5/6] Importing Platforms...")
    ws = wb["P_Platform"]

    factsheets = []
    platform_lookup = {}

    for row in ws.iter_rows(min_row=5, values_only=True):
        name = safe_str(row[1] if len(row) > 1 else None)       # col B
        plat_type = safe_str(row[2] if len(row) > 2 else None)  # col C
        friendly = safe_str(row[3] if len(row) > 3 else None)   # col D
        desc = safe_str(row[4] if len(row) > 4 else None)       # col E
        long_desc = safe_str(row[5] if len(row) > 5 else None)  # col F
        strategic = safe_str(row[6] if len(row) > 6 else None)  # col G
        comment = safe_str(row[7] if len(row) > 7 else None)    # col H
        ai = safe_str(row[8] if len(row) > 8 else None)         # col I
        owner = safe_str(row[9] if len(row) > 9 else None)      # col J

        if not name:
            continue

        guid = deterministic_guid("Platform", name)
        platform_lookup[name] = guid

        # Combine descriptions
        rich_desc_parts = []
        if desc:
            rich_desc_parts.append(desc)
        if long_desc:
            rich_desc_parts.append(long_desc)
        rich_desc = "\n\n".join(rich_desc_parts) if rich_desc_parts else None

        # Build tags
        tags = []
        if strategic:
            tags.append(f"Strategy:{strategic}")
        if ai and ai.lower() == "yes":
            tags.append("AI-enabled")

        fs = {
            "Id": guid,
            "DisplayName": name,
            "FactSheetType": "ITComponentFactSheet",
            "ITComponentType": plat_type if plat_type else "Platform",
            "ShortDescription": friendly if friendly else None,
            "RichDescription": rich_desc,
            "Responsible": owner if owner else None,
            "Tags": tags if tags else None,
        }

        if comment:
            fs["Concern"] = {
                "Header": "Platform Comment",
                "Description": comment,
                "ConcernCriticality": "Low",
            }

        factsheets.append(fs)

    log.count("Platforms (ITComponent)", len(factsheets))
    return factsheets, platform_lookup


def import_providers(wb) -> tuple:
    """
    Extract unique providers from A-Applications columns J(10) and L(12).
    Returns: (list_of_factsheets, provider_lookup_dict)
    """
    print("\n[6/6] Importing Providers...")
    ws = wb["A-Applications"]

    vendors = {}    # name -> set of types
    
    for row in ws.iter_rows(min_row=4, values_only=True):
        sw_vendor = safe_str(row[9] if len(row) > 9 else None)    # col J
        consultancy = safe_str(row[11] if len(row) > 11 else None)  # col L

        if sw_vendor:
            if sw_vendor not in vendors:
                vendors[sw_vendor] = set()
            vendors[sw_vendor].add("Software Vendor")

        if consultancy:
            if consultancy not in vendors:
                vendors[consultancy] = set()
            vendors[consultancy].add("Consultancy")

    factsheets = []
    provider_lookup = {}

    for name in sorted(vendors.keys()):
        guid = deterministic_guid("Provider", name)
        provider_lookup[name] = guid
        types = vendors[name]
        # Pick primary type
        ptype = "Software Vendor" if "Software Vendor" in types else "Consultancy"
        if len(types) > 1:
            ptype = "Software Vendor & Consultancy"

        factsheets.append({
            "Id": guid,
            "DisplayName": name,
            "FactSheetType": "ProviderFactSheet",
            "ProviderType": ptype,
        })

    log.count("Providers", len(factsheets))
    log.info(f"Sample providers: {sorted(vendors.keys())[:10]}...")
    return factsheets, provider_lookup


# ─── Hierarchy Builder ─────────────────────────────────────────────────────────

def build_hierarchy_children(factsheets: list) -> list:
    """
    For each factsheet that has children (via HierarchyParentId), add 
    HierarchyChildrenIds to the parent. This mirrors OmniGaze's hierarchy model.
    """
    id_to_fs = {fs["Id"]: fs for fs in factsheets}
    parent_children = {}  # parent_id -> [child_id, ...]

    for fs in factsheets:
        parent_id = fs.get("HierarchyParentId")
        if parent_id:
            if parent_id not in parent_children:
                parent_children[parent_id] = []
            parent_children[parent_id].append(fs["Id"])

    for parent_id, children in parent_children.items():
        if parent_id in id_to_fs:
            id_to_fs[parent_id]["HierarchyChildrenIds"] = children

    return factsheets


# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Kamstrup Phase 2 Master Data Import")
    parser.add_argument("--excel", default=DEFAULT_EXCEL, help="Path to Excel file")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Output directory")
    args = parser.parse_args()

    print(f"Kamstrup Phase 2 — Master Data Import")
    print(f"Excel: {args.excel}")
    print(f"Output: {args.output}")
    print(f"Started: {datetime.now().isoformat()}")
    print("=" * 60)

    os.makedirs(args.output, exist_ok=True)

    # Load workbook
    print("\nLoading Excel workbook...")
    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)
    print(f"Loaded. Sheets: {wb.sheetnames}")

    # 1. Value Streams
    vs_lookup = import_value_streams(wb)
    vs_list = list(vs_lookup.values())

    # 2. Business Capabilities
    bc_list, bc_id_lookup = import_business_capabilities(wb, vs_lookup)

    # 3. Organizations
    org_list, org_lookup = import_organizations(wb)

    # 4. Processes
    proc_list, proc_lookup = import_processes(wb, vs_lookup)

    # 5. Platforms
    plat_list, plat_lookup = import_platforms(wb)

    # 6. Providers
    prov_list, prov_lookup = import_providers(wb)

    wb.close()

    # Build hierarchy children IDs
    print("\nBuilding hierarchy children references...")
    bc_list = build_hierarchy_children(bc_list)
    org_list = build_hierarchy_children(org_list)
    proc_list = build_hierarchy_children(proc_list)

    # Write output files
    print("\nWriting output files...")

    def write_json(filename, data):
        path = os.path.join(args.output, filename)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)
        print(f"  -> {filename} ({len(data)} records)")

    write_json("01-value-streams.json", vs_list)
    write_json("02-business-capabilities.json", bc_list)
    write_json("03-organizations.json", org_list)
    write_json("04-processes.json", proc_list)
    write_json("05-platforms.json", plat_list)
    write_json("06-providers.json", prov_list)

    # Write lookup dictionaries for Phase 4
    lookups = {
        "business_capability_ids": bc_id_lookup,
        "organization_ids": org_lookup,
        "process_ids": proc_lookup,
        "platform_ids": plat_lookup,
        "provider_ids": prov_lookup,
        "value_stream_ids": {name: fs["Id"] for name, fs in vs_lookup.items()},
    }
    write_json("lookups-phase4.json", lookups)

    # Write combined import file
    all_factsheets = vs_list + bc_list + org_list + proc_list + plat_list + prov_list
    write_json("all-factsheets-combined.json", all_factsheets)

    # Save log
    log.save(os.path.join(args.output, "import-log.txt"))

    print(f"\n{'='*60}")
    print("IMPORT COMPLETE — Summary:")
    for k, v in log.counts.items():
        print(f"  {k}: {v}")
    print(f"  TOTAL: {sum(log.counts.values())}")
    if log.warnings:
        print(f"  Warnings: {len(log.warnings)}")
    print(f"\nOutput directory: {args.output}")
    print(f"Finished: {datetime.now().isoformat()}")


if __name__ == "__main__":
    main()
