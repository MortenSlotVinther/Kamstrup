import openpyxl
wb = openpyxl.load_workbook(r'F:\RootContext\Kamstrup\Kamstrup_Business-Capability-Map.xlsx', read_only=True, data_only=True)

# A-Applications - check row 3 for headers and rows 4-5 for data
ws = wb['A-Applications']
print('=== A-Applications row 3 (headers?) ===')
for c in range(1, ws.max_column+1):
    val = ws.cell(3, c).value
    if val:
        print(f'  col {c}: {val}')
print()
print('Row 4 sample:')
for c in [1, 2, 3, 4, 5, 10, 12, 30, 34]:
    val = ws.cell(4, c).value
    print(f'  col {c}: {val}')
print()

# B-Business Capability - check rows 3-5
ws2 = wb['B-Business Capability']
print('=== B-Business Capability row 3 ===')
for c in range(1, ws2.max_column+1):
    val = ws2.cell(3, c).value
    if val:
        print(f'  col {c}: {val}')
print()
print('Row 4-5 sample:')
for r in [4, 5]:
    vals = []
    for c in [1,2,3,4,5,6,7,8,11]:
        v = ws2.cell(r,c).value
        vals.append(f'c{c}={v}')
    print(f'  Row {r}:', '  '.join(vals))

# C-Business Context - check rows 3-5
ws3 = wb['C-Business Context']
print()
print('=== C-Business Context row 3 ===')
for c in range(1, ws3.max_column+1):
    val = ws3.cell(3, c).value
    if val:
        print(f'  col {c}: {val}')
print()
print('Row 4-5 sample:')
for r in [4, 5]:
    vals = []
    for c in [1,2,3,4,5,6,7,8,16]:
        v = ws3.cell(r,c).value
        vals.append(f'c{c}={v}')
    print(f'  Row {r}:', '  '.join(vals))

# Check KamstrupData country code to org mapping
ws4 = wb['KamstrupData']
countries = set()
for r in range(5, min(50, ws4.max_row+1)):
    cc = ws4.cell(r, 9).value
    ba = ws4.cell(r, 10).value
    dept = ws4.cell(r, 11).value
    if cc:
        countries.add((str(cc).strip(), str(ba).strip() if ba else '', str(dept).strip() if dept else ''))
print()
print('=== KamstrupData sample country/BU/Dept combos ===')
for cc, ba, dept in sorted(list(countries))[:15]:
    print(f'  CC={cc}  BA={ba}  Dept={dept}')

wb.close()
