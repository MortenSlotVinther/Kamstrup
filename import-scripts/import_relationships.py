#!/usr/bin/env python3
"""
Phase 4: Kamstrup Relationship Import Script
=============================================
Extracts ALL binary relationships from KamstrupData + A-Applications
and produces relationship JSON files for import into OmniGaze.

Reads lookups from Phase 2/3 output (lookups-phase4.json, app_number_to_guid.json)
and extracts relationships from the Excel source.

Output files:
  - output/relationships.json          (all relationships combined)
  - output/relationship_stats.json     (statistics per type)
  - output/relationship_warnings.json  (unresolved references)

Usage:
    python import_relationships.py [--excel PATH] [--output-dir DIR]

Author: Ole (Subagent)
Date: 2026-02-08
"""

import json
import os
import sys
import argparse
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ─── Configuration ──────────────────────────────────────────────────────────

EXCEL_PATH = r"F:\RootContext\Kamstrup\Kamstrup_Business-Capability-Map.xlsx"
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")

# Country code prefix → country display name (from O-Organization sheet)
COUNTRY_CODE_MAP = {
    "DK10": "Denmark",
    "DK20": "Denmark",
    "CH10": "Switzerland",
    "ES10": "Spain",
    "FR10": "France",
    "NL10": "Netherlands",
    "NO10": "Norway",
    "SE10": "Sweden",
    "US10": "United States",
    "MY10": "Malaysia",
    "PL10": "Poland",
    "AT10": "Austria",
    "CA10": "Canada",
    "CN10": "China",
    "DE10": "Germany",
    "FI10": "Finland",
    "IT10": "Italy",
}


# ─── Logging ─────────────────────────────────────────────────────────────────

class RelLog:
    """Collects import statistics and warnings."""

    def __init__(self):
        self.entries = []
        self.warnings = []
        self.stats = {}

    def info(self, msg):
        self.entries.append(f"[INFO] {msg}")
        print(f"  [INFO] {msg}")

    def warn(self, msg):
        self.warnings.append(msg)
        print(f"  [WARN] {msg}")

    def stat(self, key, val):
        self.stats[key] = val

    def summary(self):
        print("\n" + "=" * 60)
        print("RELATIONSHIP IMPORT SUMMARY")
        print("=" * 60)
        for k, v in self.stats.items():
            print(f"  {k}: {v}")
        print(f"\n  Warnings: {len(self.warnings)}")
        total = sum(v for v in self.stats.values() if isinstance(v, int))
        print(f"  TOTAL relationships: {total}")


log = RelLog()


# ─── Helpers ─────────────────────────────────────────────────────────────────

def safe_str(value) -> str:
    """Convert cell value to trimmed string, or empty string if None."""
    if value is None:
        return ""
    return str(value).strip()


def load_lookups(output_dir: str) -> dict:
    """Load all lookup dictionaries from Phase 2/3 output."""
    lookups_path = os.path.join(output_dir, "lookups-phase4.json")
    app_path = os.path.join(output_dir, "app_number_to_guid.json")

    with open(lookups_path, "r", encoding="utf-8") as f:
        lookups = json.load(f)

    with open(app_path, "r", encoding="utf-8") as f:
        app_lookup = json.load(f)

    return {
        "capabilities": lookups["business_capability_ids"],
        "organizations": lookups["organization_ids"],
        "processes": lookups["process_ids"],
        "platforms": lookups["platform_ids"],
        "providers": lookups["provider_ids"],
        "value_streams": lookups["value_stream_ids"],
        "applications": app_lookup,
    }


def resolve_app(app_str: str, app_lookup: dict) -> str:
    """Resolve application string (like '1 - IFS ERP') to GUID."""
    app_str = safe_str(app_str)
    if not app_str:
        return None
    # Try full string first
    if app_str in app_lookup:
        return app_lookup[app_str]
    # Try just the number part
    parts = app_str.split(" - ", 1)
    if parts[0].strip() in app_lookup:
        return app_lookup[parts[0].strip()]
    return None


def resolve_org_from_kamstrup_data(country_code: str, business_area: str, dept_team: str,
                                    org_lookup: dict) -> str:
    """
    Resolve organization from KamstrupData columns.
    Country code is like "DK10", business area is the dept/team name.
    Organization lookup keys use "Country::BU::Dept" format.
    
    Strategy: try most specific first, fall back to less specific.
    """
    country_code = safe_str(country_code)
    business_area = safe_str(business_area)
    dept_team = safe_str(dept_team)

    # Map country code prefix to country name
    country_name = COUNTRY_CODE_MAP.get(country_code, "")
    if not country_name:
        # Try stripping trailing digits
        prefix = country_code[:2] if len(country_code) >= 2 else ""
        for code, name in COUNTRY_CODE_MAP.items():
            if code.startswith(prefix):
                country_name = name
                break

    if not country_name:
        return None

    # The organization lookup keys from lookups-phase4.json use formats like:
    # "Denmark::Finance::DK_Finance" (Country::BU::Dept)
    # "Denmark::Finance" (Country::BU)
    # "Denmark" (Country)
    
    # The KamstrupData has: CC=DK10, BA=DK_Finance, Dept=(empty)
    # We need to match BA to the right BU and Dept.
    
    # Strategy: The BA column often matches the "Dept" level (3-part key)
    # or the "BU" level (2-part key).
    
    # Try 3-part key: "Country::BU::BA" where BA is the dept column value
    if dept_team:
        for key, guid in org_lookup.items():
            if key.startswith(f"{country_name}::") and key.endswith(f"::{dept_team}"):
                return guid

    # Try 3-part key matching BA
    if business_area:
        for key, guid in org_lookup.items():
            if key.startswith(f"{country_name}::") and key.endswith(f"::{business_area}"):
                return guid

    # Try 2-part key: "Country::BA"
    if business_area:
        key2 = f"{country_name}::{business_area}"
        if key2 in org_lookup:
            return org_lookup[key2]

    # Fall back to country level
    if country_name in org_lookup:
        return org_lookup[country_name]

    return None


# ─── Relationship Extractors ────────────────────────────────────────────────

def extract_app_capability_relationships(ws, lookups: dict) -> list:
    """
    Task 1: App → Capability from KamstrupData.
    For each row, link Application (col 12) to BusinessCapability (col 4).
    Deduplicate same app-cap pair.
    """
    print("\n[1/9] App → Capability relationships...")
    cap_lookup = lookups["capabilities"]
    app_lookup = lookups["applications"]

    relationships = []
    seen = set()
    rows_processed = 0
    unresolved_apps = set()
    unresolved_caps = set()

    for row in ws.iter_rows(min_row=5, values_only=True):
        rows_processed += 1
        app_str = safe_str(row[11] if len(row) > 11 else None)    # col 12 (0-indexed: 11)
        cap_key = safe_str(row[3] if len(row) > 3 else None)      # col 4 (0-indexed: 3)

        if not app_str or not cap_key:
            continue

        app_guid = resolve_app(app_str, app_lookup)
        cap_guid = cap_lookup.get(cap_key)

        if not app_guid:
            unresolved_apps.add(app_str)
            continue
        if not cap_guid:
            unresolved_caps.add(cap_key)
            continue

        pair = (app_guid, cap_guid)
        if pair not in seen:
            seen.add(pair)
            relationships.append({
                "type": "AppToCapability",
                "parentId": cap_guid,
                "parentType": "BusinessCapabilityFactSheet",
                "childId": app_guid,
                "childType": "ApplicationFactSheet",
                "source": "KamstrupData",
            })

    if unresolved_apps:
        log.warn(f"App→Cap: {len(unresolved_apps)} unresolved apps: {sorted(list(unresolved_apps))[:10]}")
    if unresolved_caps:
        log.warn(f"App→Cap: {len(unresolved_caps)} unresolved caps: {sorted(list(unresolved_caps))[:10]}")

    log.info(f"App→Cap: {rows_processed} rows → {len(relationships)} unique relationships")
    log.stat("App→Capability", len(relationships))
    return relationships


def extract_app_organization_relationships(ws, lookups: dict) -> list:
    """
    Task 2: App → Organization from KamstrupData.
    For each row, link Application (col 12) to Organization (cols 9-11).
    """
    print("\n[2/9] App → Organization relationships...")
    app_lookup = lookups["applications"]
    org_lookup = lookups["organizations"]

    relationships = []
    seen = set()
    unresolved_orgs = set()

    for row in ws.iter_rows(min_row=5, values_only=True):
        app_str = safe_str(row[11] if len(row) > 11 else None)
        cc = safe_str(row[8] if len(row) > 8 else None)      # col 9: CountryCode
        ba = safe_str(row[9] if len(row) > 9 else None)      # col 10: Business Area
        dept = safe_str(row[10] if len(row) > 10 else None)   # col 11: DepartmentTeam

        if not app_str or not cc:
            continue

        app_guid = resolve_app(app_str, app_lookup)
        if not app_guid:
            continue

        org_guid = resolve_org_from_kamstrup_data(cc, ba, dept, org_lookup)
        if not org_guid:
            org_key = f"{cc}|{ba}|{dept}"
            unresolved_orgs.add(org_key)
            continue

        pair = (app_guid, org_guid)
        if pair not in seen:
            seen.add(pair)
            relationships.append({
                "type": "AppToOrganization",
                "parentId": org_guid,
                "parentType": "OrganizationFactSheet",
                "childId": app_guid,
                "childType": "ApplicationFactSheet",
                "source": "KamstrupData",
            })

    if unresolved_orgs:
        log.warn(f"App→Org: {len(unresolved_orgs)} unresolved orgs: {sorted(list(unresolved_orgs))[:10]}")

    log.info(f"App→Org: {len(relationships)} unique relationships")
    log.stat("App→Organization", len(relationships))
    return relationships


def extract_app_process_relationships(ws, lookups: dict) -> list:
    """
    Task 3: App → Process from KamstrupData.
    For each row, link Application (col 12) to Process (col 16).
    """
    print("\n[3/9] App → Process relationships...")
    app_lookup = lookups["applications"]
    proc_lookup = lookups["processes"]

    relationships = []
    seen = set()
    unresolved_procs = set()

    for row in ws.iter_rows(min_row=5, values_only=True):
        app_str = safe_str(row[11] if len(row) > 11 else None)
        proc_key = safe_str(row[15] if len(row) > 15 else None)  # col 16 (0-indexed: 15)

        if not app_str or not proc_key:
            continue

        app_guid = resolve_app(app_str, app_lookup)
        if not app_guid:
            continue

        proc_guid = proc_lookup.get(proc_key)
        if not proc_guid:
            unresolved_procs.add(proc_key)
            continue

        pair = (app_guid, proc_guid)
        if pair not in seen:
            seen.add(pair)
            relationships.append({
                "type": "AppToProcess",
                "parentId": proc_guid,
                "parentType": "ProcessFactSheet",
                "childId": app_guid,
                "childType": "ApplicationFactSheet",
                "source": "KamstrupData",
            })

    if unresolved_procs:
        log.warn(f"App→Process: {len(unresolved_procs)} unresolved: {sorted(list(unresolved_procs))[:10]}")

    log.info(f"App→Process: {len(relationships)} unique relationships")
    log.stat("App→Process", len(relationships))
    return relationships


def extract_capability_valuestream_relationships(wb, lookups: dict) -> list:
    """
    Task 4: Capability → ValueStream from B-Business Capability.
    Column K (11) = Value Stream.
    """
    print("\n[4/9] Capability → ValueStream relationships...")
    ws = wb["B-Business Capability"]
    cap_lookup = lookups["capabilities"]
    vs_lookup = lookups["value_streams"]

    relationships = []
    seen = set()

    for row in ws.iter_rows(min_row=6, values_only=True):
        id_key = safe_str(row[1] if len(row) > 1 else None)   # col B = IdLevel123
        vs_name = safe_str(row[10] if len(row) > 10 else None)  # col K = Value Stream

        if not id_key or not vs_name or vs_name.lower() == "n/a":
            continue

        cap_guid = cap_lookup.get(id_key)
        vs_guid = vs_lookup.get(vs_name)

        if not cap_guid or not vs_guid:
            if not cap_guid:
                log.warn(f"Cap→VS: unresolved capability: {id_key}")
            if not vs_guid:
                log.warn(f"Cap→VS: unresolved value stream: {vs_name}")
            continue

        pair = (cap_guid, vs_guid)
        if pair not in seen:
            seen.add(pair)
            relationships.append({
                "type": "CapabilityToValueStream",
                "parentId": cap_guid,
                "parentType": "BusinessCapabilityFactSheet",
                "childId": vs_guid,
                "childType": "ValueStreamFactSheet",
                "source": "B-Business Capability",
            })

    log.info(f"Cap→VS: {len(relationships)} unique relationships")
    log.stat("Capability→ValueStream", len(relationships))
    return relationships


def extract_capability_organization_relationships(ws, lookups: dict) -> list:
    """
    Task 5: Capability → Organization from KamstrupData intersection.
    For each unique (Capability, Organization) pair.
    """
    print("\n[5/9] Capability → Organization relationships...")
    cap_lookup = lookups["capabilities"]
    org_lookup = lookups["organizations"]

    relationships = []
    seen = set()

    for row in ws.iter_rows(min_row=5, values_only=True):
        cap_key = safe_str(row[3] if len(row) > 3 else None)
        cc = safe_str(row[8] if len(row) > 8 else None)
        ba = safe_str(row[9] if len(row) > 9 else None)
        dept = safe_str(row[10] if len(row) > 10 else None)

        if not cap_key or not cc:
            continue

        cap_guid = cap_lookup.get(cap_key)
        if not cap_guid:
            continue

        org_guid = resolve_org_from_kamstrup_data(cc, ba, dept, org_lookup)
        if not org_guid:
            continue

        pair = (cap_guid, org_guid)
        if pair not in seen:
            seen.add(pair)
            relationships.append({
                "type": "CapabilityToOrganization",
                "parentId": org_guid,
                "parentType": "OrganizationFactSheet",
                "childId": cap_guid,
                "childType": "BusinessCapabilityFactSheet",
                "source": "KamstrupData",
            })

    log.info(f"Cap→Org: {len(relationships)} unique relationships")
    log.stat("Capability→Organization", len(relationships))
    return relationships


def extract_process_valuestream_relationships(wb, lookups: dict) -> list:
    """
    Task 6: Process → ValueStream from C-Business Context.
    Column P (16) = Value Stream.
    """
    print("\n[6/9] Process → ValueStream relationships...")
    ws = wb["C-Business Context"]
    proc_lookup = lookups["processes"]
    vs_lookup = lookups["value_streams"]

    relationships = []
    seen = set()

    for row in ws.iter_rows(min_row=3, values_only=True):
        no_process = safe_str(row[2] if len(row) > 2 else None)   # col C = NoProcess
        vs_name = safe_str(row[15] if len(row) > 15 else None)    # col P = Value Stream

        if not no_process or not vs_name or vs_name.lower() == "n/a":
            continue

        proc_guid = proc_lookup.get(no_process)
        vs_guid = vs_lookup.get(vs_name)

        if not proc_guid or not vs_guid:
            continue

        pair = (proc_guid, vs_guid)
        if pair not in seen:
            seen.add(pair)
            relationships.append({
                "type": "ProcessToValueStream",
                "parentId": proc_guid,
                "parentType": "ProcessFactSheet",
                "childId": vs_guid,
                "childType": "ValueStreamFactSheet",
                "source": "C-Business Context",
            })

    log.info(f"Process→VS: {len(relationships)} unique relationships")
    log.stat("Process→ValueStream", len(relationships))
    return relationships


def extract_app_provider_relationships(wb, lookups: dict) -> list:
    """
    Task 7: App → Provider from A-Applications.
    Col 10 = Software Vendor, Col 12 = Consultancy.
    """
    print("\n[7/9] App → Provider relationships...")
    ws = wb["A-Applications"]
    app_lookup = lookups["applications"]
    provider_lookup = lookups["providers"]

    relationships = []
    seen = set()
    unresolved_vendors = set()

    for row in ws.iter_rows(min_row=4, values_only=True):
        no_app = safe_str(row[1] if len(row) > 1 else None)      # col B = NoApplication
        vendor = safe_str(row[9] if len(row) > 9 else None)      # col J = Software Vendor
        consultancy = safe_str(row[11] if len(row) > 11 else None)  # col L = Consultancy

        if not no_app:
            continue

        app_guid = resolve_app(no_app, app_lookup)
        if not app_guid:
            continue

        # Software Vendor
        if vendor and vendor.lower() not in ("n/a", "", "?", "unknown"):
            vendor_guid = provider_lookup.get(vendor)
            if vendor_guid:
                pair = (app_guid, vendor_guid)
                if pair not in seen:
                    seen.add(pair)
                    relationships.append({
                        "type": "AppToProvider",
                        "parentId": app_guid,
                        "parentType": "ApplicationFactSheet",
                        "childId": vendor_guid,
                        "childType": "ProviderFactSheet",
                        "providerRole": "SoftwareVendor",
                        "source": "A-Applications",
                    })
            else:
                unresolved_vendors.add(vendor)

        # Consultancy (only if different from vendor)
        if consultancy and consultancy.lower() not in ("n/a", "", "?", "unknown"):
            if consultancy != vendor:
                consult_guid = provider_lookup.get(consultancy)
                if consult_guid:
                    pair = (app_guid, consult_guid)
                    if pair not in seen:
                        seen.add(pair)
                        relationships.append({
                            "type": "AppToProvider",
                            "parentId": app_guid,
                            "parentType": "ApplicationFactSheet",
                            "childId": consult_guid,
                            "childType": "ProviderFactSheet",
                            "providerRole": "Consultancy",
                            "source": "A-Applications",
                        })
                else:
                    unresolved_vendors.add(consultancy)

    if unresolved_vendors:
        log.warn(f"App→Provider: {len(unresolved_vendors)} unresolved: {sorted(list(unresolved_vendors))[:10]}")

    log.info(f"App→Provider: {len(relationships)} unique relationships")
    log.stat("App→Provider", len(relationships))
    return relationships


def extract_app_platform_relationships(wb, lookups: dict) -> list:
    """
    Task 8: App → Platform/ITComponent from A-Applications.
    Col 34 = Platform (may be comma-separated).
    """
    print("\n[8/9] App → Platform relationships...")
    ws = wb["A-Applications"]
    app_lookup = lookups["applications"]
    platform_lookup = lookups["platforms"]

    relationships = []
    seen = set()
    unresolved_platforms = set()

    for row in ws.iter_rows(min_row=4, values_only=True):
        no_app = safe_str(row[1] if len(row) > 1 else None)         # col B
        platform_str = safe_str(row[33] if len(row) > 33 else None)  # col AH (index 33)

        if not no_app or not platform_str or platform_str.lower() in ("n/a", ""):
            continue

        app_guid = resolve_app(no_app, app_lookup)
        if not app_guid:
            continue

        # Handle multi-value: comma or semicolon separated
        platforms = [p.strip() for p in platform_str.replace(";", ",").split(",") if p.strip()]

        for plat_name in platforms:
            if plat_name.lower() in ("n/a", ""):
                continue
            plat_guid = platform_lookup.get(plat_name)
            if plat_guid:
                pair = (app_guid, plat_guid)
                if pair not in seen:
                    seen.add(pair)
                    relationships.append({
                        "type": "AppToPlatform",
                        "parentId": app_guid,
                        "parentType": "ApplicationFactSheet",
                        "childId": plat_guid,
                        "childType": "ITComponentFactSheet",
                        "source": "A-Applications",
                    })
            else:
                unresolved_platforms.add(plat_name)

    if unresolved_platforms:
        log.warn(f"App→Platform: {len(unresolved_platforms)} unresolved: {sorted(list(unresolved_platforms))[:10]}")

    log.info(f"App→Platform: {len(relationships)} unique relationships")
    log.stat("App→Platform", len(relationships))
    return relationships


def extract_app_supporting_app_relationships(wb, lookups: dict) -> list:
    """
    Task 9: App → App (Supporting) from A-Applications.
    Col 30 = "Supporting applications" — bidirectional relationship.
    """
    print("\n[9/9] App → App (Supporting) relationships...")
    ws = wb["A-Applications"]
    app_lookup = lookups["applications"]

    relationships = []
    seen = set()
    unresolved_apps = set()

    for row in ws.iter_rows(min_row=4, values_only=True):
        no_app = safe_str(row[1] if len(row) > 1 else None)           # col B
        supporting = safe_str(row[29] if len(row) > 29 else None)      # col AD (index 29)

        if not no_app or not supporting or supporting.lower() in ("n/a", ""):
            continue

        app_guid = resolve_app(no_app, app_lookup)
        if not app_guid:
            continue

        # Supporting apps may be comma/semicolon separated, or contain "and"
        # Clean and split
        supporting_cleaned = supporting.replace(" and ", ",").replace(";", ",")
        parts = [p.strip() for p in supporting_cleaned.split(",") if p.strip()]

        for part in parts:
            if part.lower() in ("n/a", "", "none", "-"):
                continue

            # Try to resolve: could be "App Name" or "No - App Name"
            supp_guid = resolve_app(part, app_lookup)

            if not supp_guid:
                # Try fuzzy: search for apps containing this name
                part_lower = part.lower()
                for key, guid in app_lookup.items():
                    # Match on name portion (after " - ")
                    if " - " in key:
                        name_part = key.split(" - ", 1)[1].lower()
                        if name_part == part_lower or part_lower in name_part:
                            supp_guid = guid
                            break

            if not supp_guid:
                unresolved_apps.add(part)
                continue

            # Create bidirectional relationship (use sorted tuple to avoid duplicates)
            if app_guid != supp_guid:
                pair = tuple(sorted([app_guid, supp_guid]))
                if pair not in seen:
                    seen.add(pair)
                    relationships.append({
                        "type": "AppToApp",
                        "parentId": app_guid,
                        "parentType": "ApplicationFactSheet",
                        "childId": supp_guid,
                        "childType": "ApplicationFactSheet",
                        "relationshipKind": "Supporting",
                        "source": "A-Applications",
                    })

    if unresolved_apps:
        log.warn(f"App→App: {len(unresolved_apps)} unresolved: {sorted(list(unresolved_apps))[:15]}")

    log.info(f"App→App: {len(relationships)} unique relationships")
    log.stat("App→App (Supporting)", len(relationships))
    return relationships


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Phase 4: Kamstrup Relationship Import")
    parser.add_argument("--excel", default=EXCEL_PATH, help="Path to Excel file")
    parser.add_argument("--output-dir", default=OUTPUT_DIR, help="Output directory")
    args = parser.parse_args()

    print("=" * 60)
    print("Phase 4: Kamstrup Relationship Import")
    print("=" * 60)
    print(f"  Excel: {args.excel}")
    print(f"  Output: {args.output_dir}")
    print()

    # Load lookups
    print("Loading lookups from Phase 2/3...")
    lookups = load_lookups(args.output_dir)
    print(f"  Capabilities: {len(lookups['capabilities'])} entries")
    print(f"  Organizations: {len(lookups['organizations'])} entries")
    print(f"  Processes: {len(lookups['processes'])} entries")
    print(f"  Platforms: {len(lookups['platforms'])} entries")
    print(f"  Providers: {len(lookups['providers'])} entries")
    print(f"  Value Streams: {len(lookups['value_streams'])} entries")
    print(f"  Applications: {len(lookups['applications'])} entries")

    # Open Excel
    print(f"\nLoading Excel workbook (this may take a minute)...")
    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    ws_kamstrup = wb["KamstrupData"]

    # Extract all relationships
    all_relationships = []

    # 1. App → Capability (from KamstrupData)
    all_relationships.extend(extract_app_capability_relationships(ws_kamstrup, lookups))

    # 2. App → Organization (from KamstrupData)
    all_relationships.extend(extract_app_organization_relationships(ws_kamstrup, lookups))

    # 3. App → Process (from KamstrupData)
    all_relationships.extend(extract_app_process_relationships(ws_kamstrup, lookups))

    # 4. Capability → ValueStream (from B-Business Capability)
    all_relationships.extend(extract_capability_valuestream_relationships(wb, lookups))

    # 5. Capability → Organization (from KamstrupData)
    all_relationships.extend(extract_capability_organization_relationships(ws_kamstrup, lookups))

    # 6. Process → ValueStream (from C-Business Context)
    all_relationships.extend(extract_process_valuestream_relationships(wb, lookups))

    # 7. App → Provider (from A-Applications)
    all_relationships.extend(extract_app_provider_relationships(wb, lookups))

    # 8. App → Platform (from A-Applications)
    all_relationships.extend(extract_app_platform_relationships(wb, lookups))

    # 9. App → App Supporting (from A-Applications)
    all_relationships.extend(extract_app_supporting_app_relationships(wb, lookups))

    wb.close()

    # Write output
    os.makedirs(args.output_dir, exist_ok=True)

    # Combined relationships file
    output_path = os.path.join(args.output_dir, "relationships.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(all_relationships, f, indent=2, ensure_ascii=False)
    print(f"\n  Written {len(all_relationships)} relationships to {output_path}")

    # Per-type files
    by_type = defaultdict(list)
    for rel in all_relationships:
        by_type[rel["type"]].append(rel)

    for rel_type, rels in by_type.items():
        type_path = os.path.join(args.output_dir, f"relationships_{rel_type}.json")
        with open(type_path, "w", encoding="utf-8") as f:
            json.dump(rels, f, indent=2, ensure_ascii=False)

    # Stats
    stats_path = os.path.join(args.output_dir, "relationship_stats.json")
    with open(stats_path, "w", encoding="utf-8") as f:
        json.dump({
            "generated": datetime.now().isoformat(),
            "total_relationships": len(all_relationships),
            "by_type": {k: len(v) for k, v in by_type.items()},
            "stats": log.stats,
        }, f, indent=2, ensure_ascii=False)

    # Warnings
    if log.warnings:
        warnings_path = os.path.join(args.output_dir, "relationship_warnings.json")
        with open(warnings_path, "w", encoding="utf-8") as f:
            json.dump(log.warnings, f, indent=2, ensure_ascii=False)

    log.summary()
    print(f"\nDone! Files written to {args.output_dir}")


if __name__ == "__main__":
    main()
