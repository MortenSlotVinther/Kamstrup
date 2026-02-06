import openpyxl, json

wb = openpyxl.load_workbook(r'F:\RootContext\Kamstrup\Kamstrup_Business-Capability-Map.xlsx', data_only=True)
ws = wb['P-PlatformData']
lookups = json.load(open(r'F:\RootContext\Kamstrup\import-scripts\output\lookups-phase4.json'))
platform_lookup = lookups['platform_ids']

sheet_platforms = set()
for r in range(21, ws.max_row+1):
    p = ws.cell(row=r, column=2).value
    if p:
        sheet_platforms.add(p)

for p in sorted(sheet_platforms):
    match = platform_lookup.get(p, 'NOT FOUND')
    if match == 'NOT FOUND':
        print(f'  MISSING: "{p}"')
        for k in platform_lookup:
            if k.lower().startswith(p.lower()[:5]):
                print(f'    Possible match: "{k}"')
    else:
        print(f'  OK: "{p}"')
