import openpyxl
import json

wb = openpyxl.load_workbook(r'F:\RootContext\Kamstrup\Kamstrup_Business-Capability-Map.xlsx', read_only=True, data_only=True)

# Get all unique country codes + BU combos from KamstrupData
ws = wb['KamstrupData']
combos = set()
for r in range(5, ws.max_row+1):
    cc = str(ws.cell(r, 9).value or '').strip()
    ba = str(ws.cell(r, 10).value or '').strip()
    dept = str(ws.cell(r, 11).value or '').strip()
    if cc:
        combos.add((cc, ba, dept))

print(f'Unique (CC, BA, Dept) combos: {len(combos)}')
print()

# Check org mapping from O-Organization sheet
ws_org = wb['O-Organization']
print('=== O-Organization structure ===')
for r in range(1, 6):
    vals = []
    for c in range(1, 15):
        v = ws_org.cell(r, c).value
        if v:
            vals.append(f'c{c}={v}')
    print(f'Row {r}:', '  '.join(vals) if vals else '(empty)')

print()
# Show unique country codes
ccs = sorted(set(c[0] for c in combos))
print(f'Unique country codes: {ccs}')

# Check A-Applications supporting apps column
ws_app = wb['A-Applications']
print()
print('=== A-Applications: Supporting applications samples (col 30) ===')
count = 0
for r in range(4, ws_app.max_row+1):
    val = ws_app.cell(r, 30).value
    if val:
        no_app = ws_app.cell(r, 2).value
        print(f'  Row {r}: App={no_app}, Supporting={val}')
        count += 1
        if count >= 10:
            break
print(f'Total rows with supporting apps: {sum(1 for r in range(4, ws_app.max_row+1) if ws_app.cell(r, 30).value)}')

print()
print('=== A-Applications: Vendor/Platform samples ===')
for r in range(4, min(14, ws_app.max_row+1)):
    no_app = ws_app.cell(r, 2).value
    vendor = ws_app.cell(r, 10).value
    consult = ws_app.cell(r, 12).value
    platform = ws_app.cell(r, 34).value
    print(f'  App={no_app}  Vendor={vendor}  Consultancy={consult}  Platform={platform}')

wb.close()
